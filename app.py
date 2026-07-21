"""
JPMC Gastroenterology & Endoscopy Department
Centralized Booking & Slot Management System
---------------------------------------------
Single-source-of-truth booking system to eliminate multi-channel
overbooking across outpatient clinics, on-call doctors, ward admissions
and consultant direct-booking.

Uses only Flask + Python's built-in sqlite3 module — no ORM required,
so the only pip dependency is Flask itself (already preinstalled on
most hosts, including PythonAnywhere).

Run locally:
    pip install -r requirements.txt
    python app.py

Deploy on PythonAnywhere: see DEPLOY_PYTHONANYWHERE.md
"""

import os
import re
import sqlite3
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo
from functools import wraps

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, jsonify, flash, g
)
from werkzeug.security import generate_password_hash, check_password_hash

# ----------------------------------------------------------------------
# App configuration
# ----------------------------------------------------------------------
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DB_PATH = os.path.join(BASE_DIR, 'gastro_booking.db')

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'change-this-secret-key-in-production')

# ----------------------------------------------------------------------
# Constants
# ----------------------------------------------------------------------
ROLE_ADMIN = 'admin'
ROLE_SPECIALIST = 'specialist'
ROLE_SCHEDULER = 'scheduler'
ROLE_NURSE_MANAGER = 'nurse_manager'
ROLE_ONCALL = 'oncall_doctor'
ROLE_ENDOSCOPY_STAFF = 'endoscopy_staff'

ALL_ROLES = [ROLE_ADMIN, ROLE_SPECIALIST, ROLE_SCHEDULER, ROLE_NURSE_MANAGER, ROLE_ONCALL, ROLE_ENDOSCOPY_STAFF]
ROLE_LABELS = {
    ROLE_ADMIN: 'Admin (Head of Department)',
    ROLE_SPECIALIST: 'Specialist / Consultant',
    ROLE_SCHEDULER: 'Scheduler (Outpatient Clinic)',
    ROLE_NURSE_MANAGER: 'Nurse Manager (Endoscopy Suite)',
    ROLE_ONCALL: 'On-Call Doctor',
    ROLE_ENDOSCOPY_STAFF: 'Endoscopy Staff',
}
CAN_APPROVE = (ROLE_ADMIN, ROLE_SPECIALIST)
CAN_OVERRIDE = (ROLE_ADMIN, ROLE_SPECIALIST)
CAN_BOOK_ERCP = (ROLE_ADMIN, ROLE_SPECIALIST, ROLE_NURSE_MANAGER)
CAN_BOOK_SPECIAL = (ROLE_ADMIN, ROLE_SPECIALIST, ROLE_NURSE_MANAGER)
CAN_VIEW_STATS = (ROLE_ADMIN, ROLE_SPECIALIST, ROLE_NURSE_MANAGER)
# Endoscopy Staff has identical authority to Scheduler everywhere in the app —
# this tuple is the single source of truth for that equivalence.
SCHEDULER_LIKE_ROLES = (ROLE_SCHEDULER, ROLE_ENDOSCOPY_STAFF)

# ERCP Reporting Module — access restricted to the same roles who can book
# ERCP in the first place (privacy: this carries detailed clinical data,
# labs, and images, so it is not opened up beyond that group).
CAN_ACCESS_ERCP_REPORTS = CAN_BOOK_ERCP  # (admin, specialist, nurse_manager)
CAN_MANAGE_ENDOSCOPISTS = (ROLE_ADMIN, ROLE_SPECIALIST)

ERCP_IMAGE_SLOTS = 6
ERCP_IMAGES_DIR = os.path.join(BASE_DIR, 'ercp_images')
os.makedirs(ERCP_IMAGES_DIR, exist_ok=True)
ERCP_IMAGE_MAX_DIMENSION = 1600   # px, longest side
ERCP_IMAGE_JPEG_QUALITY = 78      # good balance of clinical clarity vs. storage

SEDATION_OPTIONS = [
    'Moderate Sedation (Conscious Sedation)',
    'Deep Sedation',
    'General Anesthesia',
    'None',
]
INDICATION_OPTIONS = [
    'Choledocholithiasis', 'Biliary stricture', 'Periampullary mass',
    'Acute/chronic pancreatitis', 'Cholangitis', 'Post-cholecystectomy bile leak',
    'Ampullary lesion', 'Other',
]
DUODENOSCOPE_ADVANCEMENT_OPTIONS = ['Easy', 'Mild difficulty', 'Moderate difficulty', 'Difficult', 'Very difficult']

# Papillary morphology is documented using the internationally recognized
# Haraldsson Endoscopic Papillary Classification (Types 1-4) — this
# replaces the old free-form list of descriptive appearance checkboxes.
# Location/context (e.g. periampullary diverticulum) and access difficulty
# are captured as separate fields, since they are not part of the
# Haraldsson classification itself.
PAPILLA_SHAPE_OPTIONS = [
    'Type 1 – Regular papilla',
    'Type 2 – Small / Flat papilla',
    'Type 3 – Protruding / Pendulous papilla',
    'Type 4 – Creased / Ridged papilla',
]
PAPILLA_LOCATION_OPTIONS = [
    'Periampullary diverticulum', 'Previous sphincterotomy',
    'Surgically altered anatomy', 'Periampullary mass', 'Other',
]
PAPILLA_ACCESS_OPTIONS = ['Without difficulty', 'Mild difficulty', 'Technically difficult']
CANNULATION_OPTIONS = [
    'Selective biliary cannulation', 'Difficult cannulation (needle-knife)',
    'Precut sphincterotomy', 'Pancreatic duct cannulation', 'Failed cannulation', 'Other',
]
CHOLANGIOGRAM_CATEGORIES = [
    ('Normal', ['Normal cholangiogram']),
    # Biliary Dilatation now uses numeric mm measurements (CBD/CHD/RHD/LHD)
    # instead of checkboxes — severity is classified automatically from the
    # entered values. Handled as a special case in ercp_report.html; this
    # empty option list is kept so the category still appears in the loop.
    ('Biliary Dilatation', []),
    ('Filling Defects', [
        'No filling defect', 'Single CBD stone', 'Multiple CBD stones',
        'Sludge', 'Blood clot', 'Worm', 'Filling defect (unspecified)',
    ]),
    ('Strictures', [
        'No stricture', 'Distal CBD stricture', 'Mid CBD stricture', 'Proximal CBD stricture',
        'Common hepatic duct stricture', 'Right hepatic duct stricture', 'Left hepatic duct stricture',
        'Hilar stricture', 'Bifurcation stricture', 'Benign stricture', 'Malignant stricture',
        'Indeterminate stricture', 'Multifocal strictures',
    ]),
    ('Obstruction', [
        'No obstruction', 'Partial biliary obstruction', 'Complete biliary obstruction',
        'Distal biliary obstruction', 'Hilar obstruction',
    ]),
    ('Tumours', [
        'Periampullary tumour', 'Ampullary tumour', 'Pancreatic head tumour', 'Cholangiocarcinoma',
    ]),
    ('Bile Leak', [
        'No bile leak', 'CBD leak', 'Common hepatic duct leak', 'Cystic duct leak',
        'Intrahepatic duct leak', 'Anastomotic leak', 'Post-operative bile leak', 'Contrast extravasation',
    ]),
    ('Stones & Residual Disease', [
        'Complete stone clearance', 'Residual stone', 'Residual sludge', 'Impacted stone',
        'Multiple residual stones',
    ]),
    ('Stents', [
        'No stent in situ', 'Patent plastic stent', 'Occluded plastic stent', 'Migrated plastic stent',
        'Patent metal stent', 'Occluded metal stent', 'Migrated metal stent', 'Multiple stents in situ',
    ]),
    ('Post-operative Anatomy', [
        'Post-cholecystectomy anatomy', 'Hepaticojejunostomy', 'Choledochojejunostomy',
        'Biliary-enteric anastomosis', 'Altered surgical anatomy',
    ]),
    ('Congenital / Anatomical Variants', [
        'Choledochal cyst', 'Mirizzi syndrome', 'Aberrant biliary anatomy', 'Biliary diverticulum',
        'Low cystic duct insertion', 'Variant biliary anatomy',
    ]),
    ('Sclerosing Cholangitis', [
        'Primary sclerosing cholangitis', 'Secondary sclerosing cholangitis',
        'Beading', 'Multifocal strictures', 'Irregular ducts',
    ]),
    ('Pancreatic Duct (if opacified)', [
        'Not opacified', 'Normal pancreatic duct', 'Pancreatic duct dilatation',
        'Pancreatic duct stricture', 'Pancreatic duct stone', 'Pancreatic duct leak',
        'Pancreatic duct disruption', 'Pancreatic duct irregularity',
    ]),
    ('Other Findings', [
        'Pneumobilia', 'Air bubbles', 'Poor contrast opacification', 'Contrast reflux into gallbladder',
        'Contrast reflux into pancreatic duct', 'Contrast retained within ducts', 'Debris within bile duct',
        'Other',
    ]),
]

# ---- Structured Biliary Stent Placement (replaces the old free-text field) ----
STENT_TYPE_OPTIONS = [
    'Plastic',
    'Fully Covered Self-Expandable Metal Stent (FCSEMS)',
    'Partially Covered Self-Expandable Metal Stent (PCSEMS)',
    'Uncovered Self-Expandable Metal Stent (USEMS)',
]
STENT_DIAMETER_OPTIONS = ['5 Fr', '7 Fr', '8.5 Fr', '10 Fr', '11.5 Fr']
STENT_LENGTH_OPTIONS = ['5 cm', '7 cm', '9 cm', '10 cm', '12 cm', '15 cm']
STENT_LOCATION_OPTIONS = [
    'CBD', 'Common Hepatic Duct', 'Right Hepatic Duct', 'Left Hepatic Duct',
    'Across Hilar Stricture', 'Hepaticojejunostomy', 'Other',
]
STENT_DEPLOYMENT_OPTIONS = ['Successful', 'Difficult', 'Failed']
STENT_DRAINAGE_OPTIONS = ['Good', 'Partial', 'None']
THERAPEUTIC_OPTIONS = [
    'Biliary sphincterotomy', 'Pancreatic sphincterotomy', 'Balloon dilation',
    'Stone extraction (balloon)', 'Stone extraction (basket)', 'Mechanical lithotripsy',
    'Plastic stent insertion', 'Metal stent insertion', 'Nasobiliary drain placement',
    'Needle-knife precut',
]

# ---- Procedure Note generator: natural-language phrase mappings ----
# These translate structured selections into flowing clinical sentences
# instead of a raw list of the options that were checked. (The
# cholangiogram/stent narrative builders — build_cholangiogram_sentence() /
# describe_stent_placement() — are untouched; these only cover the
# duodenoscope, papilla, cannulation, and therapeutic-steps narrative.)
DUODENOSCOPE_ADVANCEMENT_PHRASES = {
    'Easy': ('negotiated', 'without difficulty'),
    'Mild difficulty': ('negotiated', 'with mild difficulty'),
    'Moderate difficulty': ('advanced', 'with moderate difficulty'),
    'Difficult': ('advanced', 'with difficulty'),
    'Very difficult': ('advanced', 'with significant difficulty'),
}
HARALDSSON_SENTENCE = {
    'Type 1 – Regular papilla': 'The major papilla demonstrated a Haraldsson Type 1 (regular) morphology.',
    'Type 2 – Small / Flat papilla': 'The major papilla demonstrated a Haraldsson Type 2 (small/flat) morphology.',
    'Type 3 – Protruding / Pendulous papilla': 'The major papilla demonstrated a Haraldsson Type 3 (protruding/pendulous) morphology.',
    'Type 4 – Creased / Ridged papilla': 'The major papilla demonstrated a Haraldsson Type 4 (creased/ridged) morphology.',
}
PAPILLA_LOCATION_PHRASES = {
    'Periampullary diverticulum': 'within a periampullary diverticulum',
    'Previous sphincterotomy': 'in the setting of a previous sphincterotomy',
    'Surgically altered anatomy': 'in the setting of surgically altered anatomy',
    'Periampullary mass': 'in proximity to a periampullary mass',
    'Other': '',
}
CANNULATION_PHRASES = {
    'Selective biliary cannulation': 'Selective biliary cannulation was achieved using a standard sphincterotome.',
    'Difficult cannulation (needle-knife)': 'Selective biliary cannulation was difficult and required needle-knife precut.',
    'Precut sphincterotomy': 'Selective biliary cannulation was achieved following precut sphincterotomy.',
    'Pancreatic duct cannulation': 'Selective biliary cannulation was achieved following transpancreatic sphincterotomy.',
    'Failed cannulation': 'Selective biliary cannulation could not be achieved despite repeated attempts.',
}
THERAPEUTIC_PHRASES = {
    'Biliary sphincterotomy': 'a standard biliary sphincterotomy was performed',
    'Pancreatic sphincterotomy': 'a pancreatic sphincterotomy was performed',
    'Balloon dilation': 'balloon sphincteroplasty was performed',
    'Stone extraction (balloon)': 'stone extraction was completed using an extraction balloon',
    'Stone extraction (basket)': 'stone extraction was completed using a Dormia basket',
    'Mechanical lithotripsy': 'mechanical lithotripsy was required to fragment the stone(s) prior to extraction',
    'Nasobiliary drain placement': 'a nasobiliary drain was placed',
    # Stent insertion is already described by the dedicated Biliary Stent
    # Placement section (describe_stent_placement), and precut is already
    # folded into the cannulation narrative when it was actually required —
    # so neither is repeated here.
    'Plastic stent insertion': None,
    'Metal stent insertion': None,
    'Needle-knife precut': None,
}
BIOPSY_OPTIONS = [
    'Not taken', 'Brush cytology', 'Forceps biopsy', 'Brush cytology + forceps biopsy',
]
COMPLICATION_OPTIONS = [
    'None', 'Bleeding', 'Perforation', 'Post-ERCP pancreatitis',
    'Cholangitis', 'Cardiopulmonary event', 'Other',
]
CAN_MANAGE_HOLIDAYS = (ROLE_ADMIN, ROLE_SPECIALIST)

# Calendarific API — https://calendarific.com/api-documentation
# Set CALENDARIFIC_API_KEY as an environment variable (Web tab → Environment
# variables on PythonAnywhere), same as SECRET_KEY.
# API key: prefer config.py (created manually on the server, not committed to
# any zip/repo) and fall back to the CALENDARIFIC_API_KEY environment variable
# if config.py doesn't exist. This mirrors sync_holidays_task.py's lookup order.
try:
    from config import CALENDARIFIC_API_KEY as _CONFIG_API_KEY
except ImportError:
    _CONFIG_API_KEY = None
CALENDARIFIC_API_KEY = _CONFIG_API_KEY or os.environ.get('CALENDARIFIC_API_KEY', '')

# All "today" logic in this app should reflect Pakistan local time, not the
# server's (often UTC) system clock — otherwise "today" can flip to the wrong
# date for hours around midnight PKT. Timestamps stored in the DB
# (created_at, etc.) intentionally stay in UTC via datetime.utcnow().
PK_TZ = ZoneInfo('Asia/Karachi')


def today_pk():
    return datetime.now(PK_TZ).date()


CALENDARIFIC_COUNTRY = 'PK'

# Fixed-date Pakistan public holidays that land on the same month/day every
# year — applied as a permanent rule (see get_holiday()) so they work forever
# without needing yearly re-entry, API access, or database rows at all.
FIXED_ANNUAL_HOLIDAYS = [
    (2, 5, 'Kashmir Day'),
    (3, 23, 'Pakistan Day'),
    (5, 1, "Labour Day (International Workers' Day)"),
    (8, 14, 'Independence Day'),
    (12, 25, "Quaid-e-Azam Day / Christmas"),
]

# Manually-provided estimated dates for movable Islamic holidays (these shift
# every year and can't be computed from a fixed rule). Seeded once at startup
# so the calendar is protected even if the Calendarific sync never runs.
# Only 2026-2027 are seeded — beyond that, sync via Calendarific or add
# manually from the Holidays admin page.
SEED_MOVABLE_HOLIDAYS = [
    (date(2026, 8, 25), 'Eid Milad-un-Nabi (estimated)'),
    (date(2027, 3, 10), 'Eid-ul-Fitr (estimated)'),
    (date(2027, 3, 11), 'Eid-ul-Fitr (estimated)'),
    (date(2027, 3, 12), 'Eid-ul-Fitr (estimated)'),
    (date(2027, 5, 17), 'Eid-ul-Adha (estimated)'),
    (date(2027, 5, 18), 'Eid-ul-Adha (estimated)'),
    (date(2027, 5, 19), 'Eid-ul-Adha (estimated)'),
    (date(2027, 6, 15), 'Ashura (estimated)'),
    (date(2027, 6, 16), 'Ashura (estimated)'),
    (date(2027, 8, 15), 'Eid Milad-un-Nabi (estimated)'),
]

PROCEDURE_LABELS = {
    'upper_gi': 'Upper GI Endoscopy',
    'colonoscopy': 'Colonoscopy',
    'peg_tube': 'PEG Tube Insertion',
    'ercp': 'ERCP',
    'dilatation': 'Dilatation (special case)',
    'polypectomy': 'Polypectomy (special case)',
}
STANDARD_PROCEDURES = ('upper_gi', 'colonoscopy', 'peg_tube')
SPECIAL_PROCEDURES = ('dilatation', 'polypectomy')
ERCP_WEEKDAYS = (1, 5)  # Tuesday=1, Saturday=5 (Mon=0 ... Sun=6)

ONCALL_SLOT_COUNT = 8


# ----------------------------------------------------------------------
# Database helpers
# ----------------------------------------------------------------------
def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH, detect_types=sqlite3.PARSE_DECLTYPES)
        g.db.row_factory = sqlite3.Row
        g.db.execute('PRAGMA foreign_keys = ON')
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    dbconn = g.pop('db', None)
    if dbconn is not None:
        dbconn.close()


def init_db():
    dbconn = sqlite3.connect(DB_PATH)
    dbconn.row_factory = sqlite3.Row
    dbconn.executescript('''
        CREATE TABLE IF NOT EXISTS user (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            full_name TEXT NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL,
            is_approved INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS appointment (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            patient_name TEXT NOT NULL,
            gender TEXT NOT NULL,
            age INTEGER NOT NULL,
            phone TEXT NOT NULL,
            mrn TEXT NOT NULL,
            clinical_notes TEXT DEFAULT '',
            on_admission_hb TEXT DEFAULT '',
            platelet TEXT DEFAULT '',
            inr TEXT DEFAULT '',
            comorbs_etiology TEXT DEFAULT '',
            referral TEXT DEFAULT '',
            procedure_type TEXT NOT NULL,
            appointment_date TEXT NOT NULL,
            is_bleeding INTEGER NOT NULL DEFAULT 0,
            is_override INTEGER NOT NULL DEFAULT 0,
            no_show INTEGER NOT NULL DEFAULT 0,
            booked_by_username TEXT NOT NULL,
            booked_by_role TEXT NOT NULL,
            created_at TEXT NOT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_appt_date ON appointment(appointment_date);

        CREATE TABLE IF NOT EXISTS holiday (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            holiday_date TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            source TEXT NOT NULL DEFAULT 'manual',
            created_at TEXT NOT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_holiday_date ON holiday(holiday_date);

        -- ===================== ERCP Reporting Module =====================
        CREATE TABLE IF NOT EXISTS endoscopist (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name TEXT NOT NULL,
            title_lines TEXT NOT NULL DEFAULT '',
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS ercp_report (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            appointment_id INTEGER UNIQUE NOT NULL,
            status TEXT NOT NULL DEFAULT 'draft',
            endoscopist_id INTEGER,
            assistants TEXT NOT NULL DEFAULT '',
            technician TEXT NOT NULL DEFAULT '',
            sedation TEXT NOT NULL DEFAULT '',
            anesthesiologist TEXT NOT NULL DEFAULT '',
            indication TEXT NOT NULL DEFAULT '',
            duodenoscope_advancement TEXT NOT NULL DEFAULT '',
            papilla TEXT NOT NULL DEFAULT '',
            papilla_location TEXT NOT NULL DEFAULT '',
            papilla_access TEXT NOT NULL DEFAULT '',
            cannulation TEXT NOT NULL DEFAULT '',
            cholangiogram_findings TEXT NOT NULL DEFAULT '',
            cholangio_cbd_mm TEXT NOT NULL DEFAULT '',
            cholangio_chd_mm TEXT NOT NULL DEFAULT '',
            cholangio_rhd_mm TEXT NOT NULL DEFAULT '',
            cholangio_lhd_mm TEXT NOT NULL DEFAULT '',
            cholangio_largest_stone_mm TEXT NOT NULL DEFAULT '',
            cholangio_stone_count TEXT NOT NULL DEFAULT '',
            cholangio_stricture_length_mm TEXT NOT NULL DEFAULT '',
            therapeutic_procedures TEXT NOT NULL DEFAULT '',
            stent_details TEXT NOT NULL DEFAULT '',
            stent_placed TEXT NOT NULL DEFAULT '',
            stent_type TEXT NOT NULL DEFAULT '',
            stent_manufacturer TEXT NOT NULL DEFAULT '',
            stent_diameter TEXT NOT NULL DEFAULT '',
            stent_length TEXT NOT NULL DEFAULT '',
            stent_count TEXT NOT NULL DEFAULT '',
            stent_location TEXT NOT NULL DEFAULT '',
            stent_deployment TEXT NOT NULL DEFAULT '',
            stent_drainage TEXT NOT NULL DEFAULT '',
            biopsy TEXT NOT NULL DEFAULT '',
            complications TEXT NOT NULL DEFAULT '',
            procedure_note TEXT NOT NULL DEFAULT '',
            impression TEXT NOT NULL DEFAULT '',
            recommendations TEXT NOT NULL DEFAULT '',
            lab_total_bilirubin TEXT NOT NULL DEFAULT '',
            lab_direct_bilirubin TEXT NOT NULL DEFAULT '',
            lab_alt TEXT NOT NULL DEFAULT '',
            lab_ast TEXT NOT NULL DEFAULT '',
            lab_alp TEXT NOT NULL DEFAULT '',
            lab_ggt TEXT NOT NULL DEFAULT '',
            lab_albumin TEXT NOT NULL DEFAULT '',
            lab_hb TEXT NOT NULL DEFAULT '',
            lab_wbc TEXT NOT NULL DEFAULT '',
            lab_platelets TEXT NOT NULL DEFAULT '',
            lab_pt TEXT NOT NULL DEFAULT '',
            lab_inr TEXT NOT NULL DEFAULT '',
            lab_creatinine TEXT NOT NULL DEFAULT '',
            imaging_us TEXT NOT NULL DEFAULT '',
            imaging_ct TEXT NOT NULL DEFAULT '',
            imaging_mrcp TEXT NOT NULL DEFAULT '',
            finalized_by TEXT NOT NULL DEFAULT '',
            finalized_at TEXT NOT NULL DEFAULT '',
            unlocked_by TEXT NOT NULL DEFAULT '',
            unlocked_at TEXT NOT NULL DEFAULT '',
            created_by TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY (appointment_id) REFERENCES appointment(id),
            FOREIGN KEY (endoscopist_id) REFERENCES endoscopist(id)
        );

        CREATE TABLE IF NOT EXISTS ercp_report_image (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            report_id INTEGER NOT NULL,
            slot INTEGER NOT NULL,
            filename TEXT NOT NULL,
            uploaded_by TEXT NOT NULL,
            uploaded_at TEXT NOT NULL,
            UNIQUE(report_id, slot),
            FOREIGN KEY (report_id) REFERENCES ercp_report(id)
        );

        CREATE TABLE IF NOT EXISTS ercp_research (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            report_id INTEGER UNIQUE NOT NULL,
            fluoro_time_sec TEXT NOT NULL DEFAULT '',
            contrast_volume_ml TEXT NOT NULL DEFAULT '',
            cbd_diameter_mm TEXT NOT NULL DEFAULT '',
            stone_size_mm TEXT NOT NULL DEFAULT '',
            stone_count TEXT NOT NULL DEFAULT '',
            stone_clearance TEXT NOT NULL DEFAULT '',
            pd_findings TEXT NOT NULL DEFAULT '',
            pd_intervention TEXT NOT NULL DEFAULT '',
            device_details TEXT NOT NULL DEFAULT '',
            procedure_duration_min TEXT NOT NULL DEFAULT '',
            asa_class TEXT NOT NULL DEFAULT '',
            complication_severity TEXT NOT NULL DEFAULT '',
            disposition TEXT NOT NULL DEFAULT '',
            followup_plan TEXT NOT NULL DEFAULT '',
            updated_at TEXT NOT NULL,
            FOREIGN KEY (report_id) REFERENCES ercp_report(id)
        );
        CREATE INDEX IF NOT EXISTS idx_ercp_report_appt ON ercp_report(appointment_id);
        CREATE INDEX IF NOT EXISTS idx_ercp_image_report ON ercp_report_image(report_id);

        CREATE TABLE IF NOT EXISTS settings (
            id INTEGER PRIMARY KEY CHECK (id = 1),
            global_upper_gi_cap INTEGER NOT NULL DEFAULT 15,
            global_colono_cap INTEGER NOT NULL DEFAULT 5,
            global_peg_cap INTEGER NOT NULL DEFAULT 2,
            scheduler_upper_gi_cap INTEGER NOT NULL DEFAULT 10,
            scheduler_colono_cap INTEGER NOT NULL DEFAULT 3,
            warning_threshold INTEGER NOT NULL DEFAULT 12
        );
    ''')

    # --- Safe migration for databases created before these columns existed ---
    existing_settings_cols = {row['name'] for row in dbconn.execute('PRAGMA table_info(settings)').fetchall()}
    if 'global_peg_cap' not in existing_settings_cols:
        dbconn.execute("ALTER TABLE settings ADD COLUMN global_peg_cap INTEGER NOT NULL DEFAULT 2")

    existing_cols = {row['name'] for row in dbconn.execute('PRAGMA table_info(appointment)').fetchall()}
    for col in ('on_admission_hb', 'platelet', 'inr', 'comorbs_etiology', 'referral'):
        if col not in existing_cols:
            dbconn.execute(f"ALTER TABLE appointment ADD COLUMN {col} TEXT DEFAULT ''")
    if 'no_show' not in existing_cols:
        dbconn.execute("ALTER TABLE appointment ADD COLUMN no_show INTEGER NOT NULL DEFAULT 0")

    existing_ercp_cols = {row['name'] for row in dbconn.execute('PRAGMA table_info(ercp_report)').fetchall()}
    new_ercp_columns = [
        'anesthesiologist',
        'cholangio_cbd_mm', 'cholangio_chd_mm', 'cholangio_rhd_mm', 'cholangio_lhd_mm',
        'cholangio_largest_stone_mm', 'cholangio_stone_count', 'cholangio_stricture_length_mm',
        'stent_placed', 'stent_type', 'stent_manufacturer', 'stent_diameter', 'stent_length',
        'stent_count', 'stent_location', 'stent_deployment', 'stent_drainage',
        'duodenoscope_advancement', 'papilla_location', 'papilla_access',
    ]
    for col in new_ercp_columns:
        if col not in existing_ercp_cols:
            dbconn.execute(f"ALTER TABLE ercp_report ADD COLUMN {col} TEXT NOT NULL DEFAULT ''")

    dbconn.execute('INSERT OR IGNORE INTO settings (id) VALUES (1)')

    # Seed the manually-provided movable-holiday dates (Eid, Ashura, etc. for
    # 2026-2027). INSERT OR IGNORE so this never clobbers a synced or
    # manually-corrected entry that already exists for that date.
    for d, name in SEED_MOVABLE_HOLIDAYS:
        dbconn.execute(
            'INSERT OR IGNORE INTO holiday (holiday_date, name, source, created_at) VALUES (?,?,?,?)',
            (d.isoformat(), name, 'seed', datetime.utcnow().isoformat())
        )

    cur = dbconn.execute('SELECT COUNT(*) AS c FROM user WHERE role = ?', (ROLE_ADMIN,))
    if cur.fetchone()['c'] == 0:
        dbconn.execute(
            'INSERT INTO user (username, full_name, password_hash, role, is_approved, created_at) '
            'VALUES (?, ?, ?, ?, 1, ?)',
            ('admin', 'Head of Department', generate_password_hash('admin123'),
             ROLE_ADMIN, datetime.utcnow().isoformat())
        )
        print('Created default admin account -> username: admin | password: admin123')
        print('IMPORTANT: change this password immediately after first login.')

    dbconn.commit()
    dbconn.close()


# ----------------------------------------------------------------------
# Row -> dict helpers
# ----------------------------------------------------------------------
def user_role_label(role):
    return ROLE_LABELS.get(role, role)


def appt_to_dict(row):
    return {
        'id': row['id'],
        'patient_name': row['patient_name'],
        'gender': row['gender'],
        'age': row['age'],
        'phone': row['phone'],
        'mrn': row['mrn'],
        'clinical_notes': row['clinical_notes'],
        'on_admission_hb': row['on_admission_hb'],
        'platelet': row['platelet'],
        'inr': row['inr'],
        'comorbs_etiology': row['comorbs_etiology'],
        'referral': row['referral'],
        'procedure_type': row['procedure_type'],
        'procedure_label': PROCEDURE_LABELS.get(row['procedure_type'], row['procedure_type']),
        'appointment_date': row['appointment_date'],
        'is_bleeding': bool(row['is_bleeding']),
        'is_override': bool(row['is_override']),
        'no_show': bool(row['no_show']),
        'booked_by_username': row['booked_by_username'],
        'booked_by_role': row['booked_by_role'],
        'created_at': row['created_at'],
    }


# ----------------------------------------------------------------------
# Auth helpers
# ----------------------------------------------------------------------
def current_user():
    uid = session.get('user_id')
    if not uid:
        return None
    row = get_db().execute('SELECT * FROM user WHERE id = ?', (uid,)).fetchone()
    return row


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not current_user():
            return redirect(url_for('login', next=request.path))
        return view(*args, **kwargs)
    return wrapped


def roles_required(*roles):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            user = current_user()
            if not user:
                return redirect(url_for('login', next=request.path))
            if user['role'] not in roles:
                flash('You do not have permission to access that page.', 'error')
                return redirect(url_for('dashboard'))
            return view(*args, **kwargs)
        return wrapped
    return decorator


@app.context_processor
def inject_user():
    u = current_user()
    return {
        'current_user': u,
        'ROLE_LABELS': ROLE_LABELS,
        'PROCEDURE_LABELS': PROCEDURE_LABELS,
        'user_role_label': user_role_label,
    }


# ----------------------------------------------------------------------
# Business-logic helpers
# ----------------------------------------------------------------------
def get_settings():
    return get_db().execute('SELECT * FROM settings WHERE id = 1').fetchone()


def counts_for_date(d_iso, exclude_id=None):
    dbconn = get_db()
    rows = dbconn.execute('SELECT * FROM appointment WHERE appointment_date = ?', (d_iso,)).fetchall()
    result = {
        'upper_gi': 0, 'colonoscopy': 0, 'peg_tube': 0, 'ercp': 0,
        'dilatation': 0, 'polypectomy': 0,
        'scheduler_upper_gi': 0, 'scheduler_colono': 0,
        'regular_total': 0,
    }
    for a in rows:
        if exclude_id is not None and a['id'] == exclude_id:
            continue
        pt = a['procedure_type']
        result[pt] = result.get(pt, 0) + 1
        if pt in STANDARD_PROCEDURES:
            result['regular_total'] += 1
            # Scheduler/Endoscopy Staff share a dedicated sub-quota only for Upper GI / Colonoscopy;
            # PEG Tube's 2/day cap is shared openly by every role.
            if a['booked_by_role'] in SCHEDULER_LIKE_ROLES and pt in ('upper_gi', 'colonoscopy'):
                if pt == 'upper_gi':
                    result['scheduler_upper_gi'] += 1
                else:
                    result['scheduler_colono'] += 1
    return result


def day_status(d_iso):
    s = get_settings()
    c = counts_for_date(d_iso)
    ugi_cap = s['global_upper_gi_cap'] or 1
    col_cap = s['global_colono_cap'] or 1
    peg_cap = s['global_peg_cap'] or 1
    ugi_pct = c['upper_gi'] / ugi_cap
    col_pct = c['colonoscopy'] / col_cap
    peg_pct = c['peg_tube'] / peg_cap
    pct = max(ugi_pct, col_pct, peg_pct)
    if pct >= 1:
        status = 'red'
    elif pct >= 0.6:
        status = 'yellow'
    else:
        status = 'green'

    d_obj = datetime.strptime(d_iso, '%Y-%m-%d').date()
    holiday_name = get_holiday(d_iso)
    is_sunday = d_obj.weekday() == 6
    if holiday_name or is_sunday:
        status = 'holiday'

    return {
        'status': status,
        'counts': c,
        'caps': {
            'global_upper_gi': s['global_upper_gi_cap'],
            'global_colono': s['global_colono_cap'],
            'global_peg': s['global_peg_cap'],
        },
        'is_sunday': is_sunday,
        'holiday_name': holiday_name,
    }


def fetch_and_parse_holidays(window_start=None, window_end=None, api_key=None, country=None, timeout=15):
    """Fetch Pakistan public holidays from the Calendarific API for every year
    spanned by the given window, and return a list of (date, name) tuples.
    Filters to nationally-observed public holidays only — Calendarific also
    returns minor "observance" entries (e.g. international awareness days)
    that shouldn't block hospital bookings."""
    import requests

    key = api_key or CALENDARIFIC_API_KEY
    if not key:
        raise RuntimeError(
            'CALENDARIFIC_API_KEY is not set. Add it under the Web tab → '
            'Environment variables on PythonAnywhere (same place as SECRET_KEY), '
            'then reload the app.'
        )

    today = today_pk()
    window_start = window_start or (today - timedelta(days=30))
    window_end = window_end or (today + timedelta(days=730))
    years = sorted({window_start.year, window_end.year})

    events = []
    for year in years:
        resp = requests.get(
            'https://calendarific.com/api/v2/holidays',
            params={'api_key': key, 'country': country or CALENDARIFIC_COUNTRY, 'year': year},
            timeout=timeout,
        )
        resp.raise_for_status()
        data = resp.json()
        meta = data.get('meta', {})
        if meta.get('code') != 200:
            raise RuntimeError(f"Calendarific error: {meta.get('error_detail', 'unknown error')}")

        for h in data.get('response', {}).get('holidays', []):
            types = [t.lower() for t in h.get('type', [])]
            if not any('national' in t or 'public' in t for t in types):
                continue  # skip observances / minor days that aren't actual off-days
            iso = h.get('date', {}).get('iso', '')
            try:
                d = datetime.strptime(iso[:10], '%Y-%m-%d').date()
            except ValueError:
                continue
            name = h.get('name', '').strip()
            if name:
                events.append((d, name))
    return events


def sync_holidays_into_db(dbconn, events, window_days_past=30, window_days_future=730):
    """Upsert a list of (date, name) tuples into the holiday table, scoped to
    a reasonable window around today so the table doesn't fill with old noise.
    Manually-added holidays (source='manual') are never touched/overwritten."""
    today = today_pk()
    window_start = today - timedelta(days=window_days_past)
    window_end = today + timedelta(days=window_days_future)
    count = 0
    for d, name in events:
        if window_start <= d <= window_end:
            dbconn.execute(
                'INSERT INTO holiday (holiday_date, name, source, created_at) VALUES (?,?,?,?) '
                "ON CONFLICT(holiday_date) DO UPDATE SET name=excluded.name, source='auto_sync' "
                "WHERE holiday.source = 'auto_sync'",
                (d.isoformat(), name, 'auto_sync', datetime.utcnow().isoformat())
            )
            count += 1
    dbconn.commit()
    return count


def get_holiday(d_iso):
    """Return the holiday name for a given ISO date string, or None.
    Checks the DB table first (covers synced/seeded/manually-added holidays),
    then falls back to the permanent fixed-date rule set — national holidays
    that land on the same month/day every year, so they work indefinitely
    with no database rows or API access needed."""
    row = get_db().execute('SELECT name FROM holiday WHERE holiday_date = ?', (d_iso,)).fetchone()
    if row:
        return row['name']
    d = datetime.strptime(d_iso, '%Y-%m-%d').date()
    for month, day, name in FIXED_ANNUAL_HOLIDAYS:
        if d.month == month and d.day == day:
            return name
    return None


def is_time_locked(row):
    """True once the appointment day has arrived OR 48 hours have passed since
    the booking was created — the cutoff used for Scheduler edit/delete limits
    and the Admin-only deletion window."""
    try:
        appt_date = datetime.strptime(row['appointment_date'], '%Y-%m-%d').date()
        day_arrived = today_pk() >= appt_date
    except (ValueError, TypeError):
        day_arrived = False

    try:
        created_dt = datetime.fromisoformat(row['created_at'])
        hours_since = (datetime.utcnow() - created_dt).total_seconds() / 3600
        past_48h = hours_since >= 48
    except (ValueError, TypeError):
        past_48h = False

    return day_arrived or past_48h


def parse_numeric(text):
    """Best-effort extraction of the first number in a free-text lab value
    like '11.2 g/dL' or '180' or 'N/A'. Returns None if nothing parseable."""
    if not text:
        return None
    match = re.search(r'[-+]?\d*\.\d+|[-+]?\d+', text)
    if not match:
        return None
    try:
        return float(match.group())
    except ValueError:
        return None


def compute_statistics(start_d, end_d):
    """Aggregate medical statistics for all bookings within an inclusive date range."""
    s = get_settings()
    rows = get_db().execute(
        'SELECT * FROM appointment WHERE appointment_date BETWEEN ? AND ?',
        (start_d.isoformat(), end_d.isoformat())
    ).fetchall()

    by_procedure = {k: 0 for k in PROCEDURE_LABELS}
    genders = {}
    ages = []
    hb_vals, plt_vals, inr_vals = [], [], []
    bleeding_count = 0
    override_count = 0
    booked_by_role = {}
    lead_times = []

    # Full daily grid across the whole range (including zero-booking days) —
    # needed for accurate cap-utilization averages and the timeline chart.
    daily = {}
    d = start_d
    while d <= end_d:
        daily[d.isoformat()] = {'upper_gi': 0, 'colonoscopy': 0, 'peg_tube': 0, 'ercp': 0, 'special': 0}
        d += timedelta(days=1)

    for r in rows:
        pt = r['procedure_type']
        by_procedure[pt] = by_procedure.get(pt, 0) + 1
        g = r['gender'] or 'Unspecified'
        genders[g] = genders.get(g, 0) + 1
        ages.append(r['age'])

        hb = parse_numeric(r['on_admission_hb'])
        if hb is not None:
            hb_vals.append(hb)
        plt = parse_numeric(r['platelet'])
        if plt is not None:
            plt_vals.append(plt)
        inr = parse_numeric(r['inr'])
        if inr is not None:
            inr_vals.append(inr)

        if r['is_bleeding']:
            bleeding_count += 1
        if r['is_override']:
            override_count += 1

        role = r['booked_by_role']
        booked_by_role[role] = booked_by_role.get(role, 0) + 1

        d_iso = r['appointment_date']
        bucket = 'special' if pt in SPECIAL_PROCEDURES else (pt if pt in ('upper_gi', 'colonoscopy', 'peg_tube', 'ercp') else None)
        if d_iso in daily and bucket:
            daily[d_iso][bucket] += 1

        try:
            created_date = datetime.fromisoformat(r['created_at']).date()
            appt_date = datetime.strptime(r['appointment_date'], '%Y-%m-%d').date()
            lead_times.append((appt_date - created_date).days)
        except (ValueError, TypeError):
            pass

    def avg(values):
        return round(sum(values) / len(values), 2) if values else None

    total = len(rows)
    span_days = (end_d - start_d).days + 1
    avg_bookings_per_day = round(total / span_days, 2) if span_days else None

    # Average daily cap utilization across the whole period (0 on empty days too)
    utilization_vals = []
    if s['global_upper_gi_cap']:
        utilization_vals += [c['upper_gi'] / s['global_upper_gi_cap'] for c in daily.values()]
    if s['global_colono_cap']:
        utilization_vals += [c['colonoscopy'] / s['global_colono_cap'] for c in daily.values()]
    if s['global_peg_cap']:
        utilization_vals += [c['peg_tube'] / s['global_peg_cap'] for c in daily.values()]
    avg_utilization_pct = round((sum(utilization_vals) / len(utilization_vals)) * 100, 1) if utilization_vals else None

    # Bucket the daily grid into day/week/month points depending on range length
    if span_days <= 45:
        bucket_unit = 'day'
    elif span_days <= 180:
        bucket_unit = 'week'
    else:
        bucket_unit = 'month'
    timeline = bucket_timeline(start_d, end_d, daily, bucket_unit)

    return {
        'total': total,
        'span_days': span_days,
        'avg_bookings_per_day': avg_bookings_per_day,
        'by_procedure': by_procedure,
        'genders': genders,
        'avg_age': avg(ages),
        'avg_hb': avg(hb_vals), 'hb_n': len(hb_vals),
        'avg_platelet': avg(plt_vals), 'platelet_n': len(plt_vals),
        'avg_inr': avg(inr_vals), 'inr_n': len(inr_vals),
        'bleeding_count': bleeding_count,
        'override_count': override_count,
        'booked_by_role': {user_role_label(k): v for k, v in booked_by_role.items()},
        'avg_lead_time_days': avg(lead_times),
        'avg_utilization_pct': avg_utilization_pct,
        'timeline': timeline,
        'bucket_unit': bucket_unit,
    }


def bucket_timeline(start_d, end_d, daily, bucket_unit):
    """Group the per-day counts into day/week/month buckets for the trend chart."""
    buckets = {}
    order = []
    d = start_d
    while d <= end_d:
        d_iso = d.isoformat()
        counts = daily.get(d_iso, {'upper_gi': 0, 'colonoscopy': 0, 'peg_tube': 0, 'ercp': 0, 'special': 0})
        if bucket_unit == 'day':
            key = d_iso
            label = d.strftime('%b %d')
        elif bucket_unit == 'week':
            week_start = d - timedelta(days=d.weekday())
            key = week_start.isoformat()
            label = 'Wk of ' + week_start.strftime('%b %d')
        else:
            key = f'{d.year}-{d.month:02d}'
            label = d.strftime('%b %Y')
        if key not in buckets:
            buckets[key] = {'label': label, 'upper_gi': 0, 'colonoscopy': 0, 'peg_tube': 0, 'ercp': 0, 'special': 0}
            order.append(key)
        for k in ('upper_gi', 'colonoscopy', 'peg_tube', 'ercp', 'special'):
            buckets[key][k] += counts[k]
        d += timedelta(days=1)
    return [buckets[k] for k in order]


# ===================== ERCP Reporting Module — helpers =====================

# ---- Cholangiogram structured-finding lookup tables (used to turn checkbox
# selections + numeric measurements into a natural-language sentence, and to
# keep those items out of the generic "leftover findings" list so nothing is
# described twice in the auto-generated note). ----
STRICTURE_LOCATION_TEXT = {
    'Distal CBD stricture': 'distal common bile duct',
    'Mid CBD stricture': 'mid common bile duct',
    'Proximal CBD stricture': 'proximal common bile duct',
    'Common hepatic duct stricture': 'common hepatic duct',
    'Right hepatic duct stricture': 'right hepatic duct',
    'Left hepatic duct stricture': 'left hepatic duct',
    'Hilar stricture': 'hilar',
    'Bifurcation stricture': 'biliary bifurcation',
    'Multifocal strictures': 'multifocal',
}
STRICTURE_NATURE_TEXT = {
    'Benign stricture': 'benign',
    'Malignant stricture': 'malignant',
    'Indeterminate stricture': 'indeterminate',
}
FILLING_DEFECT_TEXT = {
    'Single CBD stone': 'a single CBD stone',
    'Multiple CBD stones': 'multiple CBD stones',
    'Sludge': 'biliary sludge',
    'Blood clot': 'a blood clot',
    'Worm': 'a biliary worm',
    'Filling defect (unspecified)': 'an unspecified filling defect',
}

# ---- Structured Biliary Stent Placement lookup tables ----
STENT_TYPE_SHORT_TEXT = {
    'Plastic': 'plastic',
    'Fully Covered Self-Expandable Metal Stent (FCSEMS)': 'fully covered self-expandable metal',
    'Partially Covered Self-Expandable Metal Stent (PCSEMS)': 'partially covered self-expandable metal',
    'Uncovered Self-Expandable Metal Stent (USEMS)': 'uncovered self-expandable metal',
}
STENT_DEPLOYMENT_TEXT = {
    'Successful': 'successfully deployed',
    'Difficult': 'deployed with technical difficulty',
    'Failed': 'unsuccessfully attempted',
}
STENT_LOCATION_TEXT = {
    'CBD': 'in the common bile duct',
    'Common Hepatic Duct': 'in the common hepatic duct',
    'Right Hepatic Duct': 'in the right hepatic duct',
    'Left Hepatic Duct': 'in the left hepatic duct',
    'Across Hilar Stricture': 'across the hilar stricture',
    'Hepaticojejunostomy': 'at the hepaticojejunostomy',
    'Other': '',
}
STENT_DRAINAGE_TEXT = {
    'Good': 'with good bile drainage',
    'Partial': 'with partial bile drainage',
    'None': 'with no bile drainage observed',
}


def classify_dilatation_mm(mm):
    """Automatic severity classification for a biliary duct measurement, in
    millimetres. Documentation aid only — always reviewable/editable by the
    endoscopist, same as the rest of the auto-generated note."""
    if mm is None:
        return None
    if mm >= 15:
        return 'Marked'
    if mm >= 10:
        return 'Moderate'
    if mm >= 7:
        return 'Mild'
    return 'No dilatation'


def describe_biliary_dilatation(fields):
    """Natural-language fragment for the most dilated of the four measured
    ducts (CBD / CHD / right & left hepatic ducts), or '' if none entered."""
    ducts = [
        ('common bile duct', fields.get('cholangio_cbd_mm')),
        ('common hepatic duct', fields.get('cholangio_chd_mm')),
        ('right hepatic duct', fields.get('cholangio_rhd_mm')),
        ('left hepatic duct', fields.get('cholangio_lhd_mm')),
    ]
    measured = [(label, parse_numeric(raw)) for label, raw in ducts]
    measured = [(label, mm) for label, mm in measured if mm is not None]
    if not measured:
        return ''
    label, mm = max(measured, key=lambda pair: pair[1])
    mm_str = f'{mm:g}'
    if classify_dilatation_mm(mm) == 'No dilatation':
        return f'a {label} measuring {mm_str} mm with no significant dilatation'
    return f'a dilated {label} measuring {mm_str} mm'


def describe_stricture(selected_findings, length_mm_raw):
    """Natural-language fragment combining stricture nature + location +
    (optional) length, built from the checked Strictures-category findings."""
    natures = [STRICTURE_NATURE_TEXT[v] for v in selected_findings if v in STRICTURE_NATURE_TEXT]
    locations = [STRICTURE_LOCATION_TEXT[v] for v in selected_findings if v in STRICTURE_LOCATION_TEXT]
    if not natures and not locations:
        return ''
    bits = []
    if natures:
        bits.append('/'.join(natures))
    if locations:
        bits.append('/'.join(locations))
    descriptor = ' '.join(bits)
    length_mm = parse_numeric(length_mm_raw)
    length_text = f' measuring approximately {length_mm:g} mm in length' if length_mm is not None else ''
    return f'a {descriptor} stricture{length_text}'


def describe_filling_defects(selected_findings, largest_mm_raw, count_raw):
    """Natural-language fragment for filling defects, preferring the numeric
    count/size if entered, falling back to the checked defect types."""
    items = [FILLING_DEFECT_TEXT[v] for v in selected_findings if v in FILLING_DEFECT_TEXT]
    if not items:
        return ''
    count = parse_numeric(count_raw)
    largest = parse_numeric(largest_mm_raw)
    if count is not None and count >= 1:
        count_int = int(count)
        base = f"{count_int} filling defect{'s' if count_int != 1 else ''}"
    elif len(items) == 1:
        base = items[0]
    else:
        base = ', '.join(items[:-1]) + f', and {items[-1]}'
    if largest is not None:
        base += f', the largest measuring {largest:g} mm'
    return base


def build_cholangiogram_sentence(fields):
    """Compose a natural-language cholangiogram description from the
    structured dilatation/stricture/filling-defect fields plus any other
    checked findings, without repeating information across sentences.
    Accepts a plain dict (used both for the live 'Generate Note' payload and,
    via dict(report), for the printed-report summary)."""
    raw = fields.get('cholangiogram_findings') or ''
    selected = [v.strip() for v in raw.split(',') if v.strip()]

    has_measurement = any(
        parse_numeric(fields.get(k)) is not None for k in (
            'cholangio_cbd_mm', 'cholangio_chd_mm', 'cholangio_rhd_mm', 'cholangio_lhd_mm',
            'cholangio_largest_stone_mm', 'cholangio_stone_count', 'cholangio_stricture_length_mm',
        )
    )
    if selected == ['Normal cholangiogram'] and not has_measurement:
        return 'Cholangiogram was normal.'

    dilation_frag = describe_biliary_dilatation(fields)
    stricture_frag = describe_stricture(selected, fields.get('cholangio_stricture_length_mm'))
    filling_frag = describe_filling_defects(
        selected, fields.get('cholangio_largest_stone_mm'), fields.get('cholangio_stone_count')
    )
    structured_bits = [b for b in (dilation_frag, stricture_frag, filling_frag) if b]

    leftover = [
        v for v in selected
        if v not in STRICTURE_LOCATION_TEXT and v not in STRICTURE_NATURE_TEXT
        and v not in FILLING_DEFECT_TEXT
        and v not in ('Normal cholangiogram', 'No stricture', 'No filling defect')
    ]

    sentences = []
    if structured_bits:
        if len(structured_bits) == 1:
            joined = structured_bits[0]
        elif len(structured_bits) == 2:
            joined = f'{structured_bits[0]} with {structured_bits[1]}'
        else:
            joined = f'{structured_bits[0]}, with {structured_bits[1]}, and {structured_bits[2]}'
        sentences.append(f'Cholangiography demonstrated {joined}.')
    if leftover:
        prefix = 'Cholangiogram also demonstrated' if sentences else 'Cholangiogram demonstrated'
        sentences.append(f'{prefix}: {", ".join(leftover)}.')
    return ' '.join(sentences)


def describe_stent_placement(fields):
    """Natural-language biliary stent sentence built from the structured
    Biliary Stent Placement fields. Falls back to the legacy free-text
    'stent_details' value for reports created before this section existed
    (that column is preserved untouched — never overwritten by the new form).
    Accepts a plain dict (payload dict for live note generation, or
    dict(report) for the printed-report summary)."""
    placed = (fields.get('stent_placed') or '').strip()
    if not placed:
        return (fields.get('stent_details') or '').strip()
    if placed == 'No':
        return 'No biliary stent was placed.'

    stype = (fields.get('stent_type') or '').strip()
    manufacturer = (fields.get('stent_manufacturer') or '').strip()
    diameter = (fields.get('stent_diameter') or '').strip()
    length = (fields.get('stent_length') or '').strip()
    location = (fields.get('stent_location') or '').strip()
    deployment = (fields.get('stent_deployment') or '').strip()
    drainage = (fields.get('stent_drainage') or '').strip()

    count_n = parse_numeric(fields.get('stent_count'))
    count_n = int(count_n) if count_n and count_n >= 1 else 1
    plural = count_n > 1

    type_adj = STENT_TYPE_SHORT_TEXT.get(stype, stype.lower())
    size_bit = ' × '.join(x for x in (diameter, length) if x)
    noun = 'biliary stents' if plural else 'biliary stent'
    descriptor = ' '.join(x for x in (size_bit, type_adj) if x)

    subject = f'{count_n} {descriptor} {noun}' if plural else f'A {descriptor} {noun}'
    subject = ' '.join(subject.split())
    if subject and subject[0].islower():
        subject = subject[0].upper() + subject[1:]

    verb = 'were' if plural else 'was'
    deployment_phrase = STENT_DEPLOYMENT_TEXT.get(deployment, 'deployed')
    sentence = f'{subject} {verb} {deployment_phrase}'

    location_phrase = STENT_LOCATION_TEXT.get(location, '')
    if location_phrase:
        sentence += f' {location_phrase}'

    if drainage and deployment != 'Failed':
        drainage_phrase = STENT_DRAINAGE_TEXT.get(drainage, '')
        if drainage_phrase:
            sentence += f', {drainage_phrase}'

    sentence += '.'
    if manufacturer:
        sentence += f' Manufacturer: {manufacturer}.'
    return sentence


def describe_papilla_summary(fields):
    """Compact 'Shape; Location; Access' summary for the printed report's
    Papilla row (kept separate from build_papilla_sentences, which produces
    the full narrative sentences used in the procedure note)."""
    bits = [
        (fields.get('papilla') or '').strip(),
        (fields.get('papilla_location') or '').strip(),
        (fields.get('papilla_access') or '').strip(),
    ]
    return '; '.join(b for b in bits if b)


def build_duodenoscope_sentence(advancement):
    """Describe duodenoscope advancement to D2 in the wording of an
    experienced endoscopist rather than a flat, always-identical sentence."""
    advancement = (advancement or '').strip()
    if not advancement:
        return ''
    verb, phrase = DUODENOSCOPE_ADVANCEMENT_PHRASES.get(advancement, ('advanced', ''))
    if phrase:
        return f'The duodenoscope was {verb} to the second part of the duodenum {phrase}.'
    return f'The duodenoscope was {verb} to the second part of the duodenum.'


def build_papilla_sentences(papilla_shape, papilla_location, papilla_access):
    """Describe papillary access/location, then morphology (Haraldsson
    classification), as two natural sentences. Returns [] if nothing about
    the papilla was documented."""
    papilla_shape = (papilla_shape or '').strip()
    papilla_location = (papilla_location or '').strip()
    papilla_access = (papilla_access or '').strip()
    if not (papilla_shape or papilla_location or papilla_access):
        return []

    sentences = []
    loc_phrase = PAPILLA_LOCATION_PHRASES.get(papilla_location, '')
    if papilla_access == 'Technically difficult':
        sentence = 'Access to the papilla was technically difficult'
        if loc_phrase:
            sentence += f', the papilla being located {loc_phrase}'
        sentences.append(sentence + '.')
    else:
        bits = []
        if loc_phrase:
            bits.append(loc_phrase)
        if papilla_access == 'Mild difficulty':
            bits.append('with mild difficulty')
        if not bits:
            bits.append('without difficulty')
        sentences.append(f'The major papilla was identified {", ".join(bits)}.')

    if papilla_shape:
        sentences.append(HARALDSSON_SENTENCE.get(
            papilla_shape, f'The major papilla demonstrated a {papilla_shape} morphology.'
        ))

    return sentences


def build_cannulation_sentence(cannulation):
    """Describe how biliary access was obtained, reflecting what was
    actually performed rather than a single generic sentence."""
    cannulation = (cannulation or '').strip()
    if not cannulation:
        return ''
    return CANNULATION_PHRASES.get(cannulation, f'{cannulation} was achieved.')


def build_therapeutic_paragraph(therapeutic_str):
    """Turn the selected therapeutic-step checkboxes into a procedural
    paragraph describing the sequence of interventions, rather than a raw
    comma-separated list."""
    items = [t.strip() for t in (therapeutic_str or '').split(',') if t.strip()]
    if not items:
        return ''

    fragments = []
    extraction_used = False
    for item in items:
        phrase = THERAPEUTIC_PHRASES.get(item)
        if not phrase:
            continue
        fragments.append(phrase)
        if item.startswith('Stone extraction'):
            extraction_used = True

    if not fragments:
        return ''
    if extraction_used:
        fragments[-1] += ', achieving satisfactory ductal clearance'

    sentence = fragments[0]
    for frag in fragments[1:]:
        sentence += f'; {frag}'
    return sentence[0].upper() + sentence[1:] + '.'


def generate_procedure_note(fields):
    """Build a draft procedure-note narrative from structured ERCP fields,
    written the way an experienced therapeutic endoscopist would dictate it
    — flowing clinical sentences grouped by theme, not a checklist read-out.
    Administrative details already captured elsewhere in the report
    (sedation, indication, anesthesiologist/assistants/technician) are
    intentionally NOT restated here. Always meant to be reviewed/edited
    before finalizing.

    The cholangiogram narrative (build_cholangiogram_sentence) and the
    biliary stent narrative (describe_stent_placement) are unchanged —
    this function only restructures the surrounding sections."""
    duodenoscope_advancement = fields.get('duodenoscope_advancement') or ''
    papilla_shape = fields.get('papilla') or ''
    papilla_location = fields.get('papilla_location') or ''
    papilla_access = fields.get('papilla_access') or ''
    cannulation = fields.get('cannulation') or ''
    therapeutic = fields.get('therapeutic_procedures') or ''
    biopsy = fields.get('biopsy') or ''
    complications = fields.get('complications') or ''

    parts = []

    duodenoscope_sentence = build_duodenoscope_sentence(duodenoscope_advancement)
    if duodenoscope_sentence:
        parts.append(duodenoscope_sentence)

    parts.extend(build_papilla_sentences(papilla_shape, papilla_location, papilla_access))

    cannulation_sentence = build_cannulation_sentence(cannulation)
    if cannulation_sentence:
        parts.append(cannulation_sentence)

    cholangio_sentence = build_cholangiogram_sentence(fields)
    if cholangio_sentence:
        parts.append(cholangio_sentence)

    therapeutic_paragraph = build_therapeutic_paragraph(therapeutic)
    if therapeutic_paragraph:
        parts.append(therapeutic_paragraph)

    stent_sentence = describe_stent_placement(fields)
    if stent_sentence:
        parts.append(stent_sentence)

    if biopsy and biopsy.lower() != 'not taken':
        parts.append(f'{biopsy} was obtained and sent for histopathology.')

    if complications and complications.lower() != 'none':
        parts.append(f'The following complication(s) were noted during the procedure: {complications}.')
    else:
        parts.append('The patient tolerated the procedure well, and no immediate procedure-related complications were observed.')

    return ' '.join(parts)



def compress_ercp_image(file_storage, dest_path):
    """Resize/compress an uploaded image and save it as a JPEG at dest_path."""
    from PIL import Image
    img = Image.open(file_storage)
    img = img.convert('RGB')
    img.thumbnail((ERCP_IMAGE_MAX_DIMENSION, ERCP_IMAGE_MAX_DIMENSION), Image.LANCZOS)
    img.save(dest_path, 'JPEG', quality=ERCP_IMAGE_JPEG_QUALITY, optimize=True)


def generate_qr_data_uri(url):
    """Return a base64 data: URI for a QR code pointing at the given URL,
    or None if the qrcode library isn't available — the print page degrades
    gracefully (no QR box) rather than failing outright."""
    try:
        import qrcode
        import io
        import base64
        qr = qrcode.QRCode(box_size=6, border=2)
        qr.add_data(url)
        qr.make(fit=True)
        img = qr.make_image(fill_color='black', back_color='white')
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        encoded = base64.b64encode(buf.getvalue()).decode('ascii')
        return f'data:image/png;base64,{encoded}'
    except Exception:
        return None


def get_or_create_ercp_report(appointment_id, username):
    """One report per appointment, created lazily on first open."""
    dbconn = get_db()
    row = dbconn.execute('SELECT * FROM ercp_report WHERE appointment_id = ?', (appointment_id,)).fetchone()
    if row:
        return row
    now = datetime.utcnow().isoformat()
    dbconn.execute(
        'INSERT INTO ercp_report (appointment_id, status, created_by, created_at, updated_at) '
        'VALUES (?, ?, ?, ?, ?)',
        (appointment_id, 'draft', username, now, now)
    )
    dbconn.execute(
        'INSERT INTO ercp_research (report_id, updated_at) '
        'SELECT id, ? FROM ercp_report WHERE appointment_id = ?',
        (now, appointment_id)
    )
    dbconn.commit()
    return dbconn.execute('SELECT * FROM ercp_report WHERE appointment_id = ?', (appointment_id,)).fetchone()


def validate_booking(user, payload, exclude_id=None):
    """Validate a booking (or edit) request against all business rules.
    exclude_id: when editing an existing appointment, its own id is excluded
    from cap/quota counts so it doesn't collide with itself.
    Returns (ok: bool, error_message: str or None, extra: dict)
    """
    s = get_settings()
    procedure_type = payload.get('procedure_type')
    appt_date_str = payload.get('appointment_date')
    is_override = bool(payload.get('is_override'))

    if procedure_type not in PROCEDURE_LABELS:
        return False, 'Invalid procedure type.', {}

    try:
        d = datetime.strptime(appt_date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return False, 'Invalid or missing appointment date.', {}

    d_iso_check = d.isoformat()
    holiday_name = get_holiday(d_iso_check)
    if (d.weekday() == 6 or holiday_name) and user['role'] not in CAN_OVERRIDE:
        if holiday_name:
            return False, (
                f'{d_iso_check} is a public holiday ({holiday_name}). '
                f'Only Admin or Specialist accounts can book on this date.'
            ), {}
        return False, (
            f'{d_iso_check} is a Sunday. Only Admin or Specialist accounts can book on this date.'
        ), {}

    # MRN is optional; everything else about the patient is required
    for field in ('patient_name', 'gender', 'age', 'phone'):
        if not payload.get(field):
            return False, f'"{field.replace("_", " ").title()}" is required.', {}

    try:
        age = int(payload.get('age'))
        if age <= 0 or age > 130:
            return False, 'Age must be a realistic integer.', {}
    except (ValueError, TypeError):
        return False, 'Age must be a valid integer.', {}

    d_iso = d.isoformat()

    # ERCP: Tuesdays and Saturdays only, restricted roles.
    # Patient data can be entered on any day of the week - it's only the
    # appointment_date itself that must fall on an ERCP list day.
    if procedure_type == 'ercp':
        if d.weekday() not in ERCP_WEEKDAYS:
            return False, 'ERCP can only be scheduled on Tuesdays or Saturdays.', {}
        if user['role'] not in CAN_BOOK_ERCP:
            return False, 'Only Admin, Specialist or Nurse Manager may book ERCP cases.', {}
        return True, None, {}

    # Special diagnostic/therapeutic cases (do not count toward 15/5 cap)
    if procedure_type in SPECIAL_PROCEDURES:
        if user['role'] not in CAN_BOOK_SPECIAL:
            return False, 'Only Nurse Manager, Specialist or Admin may book special cases.', {}
        c = counts_for_date(d_iso, exclude_id=exclude_id)
        warn = c['regular_total'] > s['warning_threshold']
        return True, None, {'warning': warn}

    # Standard procedures: upper_gi / colonoscopy / peg_tube
    if procedure_type in STANDARD_PROCEDURES:
        c = counts_for_date(d_iso, exclude_id=exclude_id)
        cap_field_map = {
            'upper_gi': 'global_upper_gi_cap',
            'colonoscopy': 'global_colono_cap',
            'peg_tube': 'global_peg_cap',
        }
        cap = s[cap_field_map[procedure_type]]
        current_count = c[procedure_type]

        # Admin/Specialist manual override for emergencies bypasses global cap
        if is_override:
            if user['role'] not in CAN_OVERRIDE:
                return False, 'Only Admin or Specialist accounts can override capacity limits.', {}
            return True, None, {'override_used': True}

        if current_count >= cap:
            proc_label = PROCEDURE_LABELS[procedure_type]
            return False, (
                f'Hard cap reached: {proc_label} is fully booked for {d_iso} '
                f'({current_count}/{cap}). Only Admin/Specialist can override for an emergency.'
            ), {}

        # Scheduler-specific quota (preserves remaining slots for on-call/inpatients).
        # PEG Tube has no separate Scheduler sub-quota — its 2/day cap is shared openly.
        if user['role'] in SCHEDULER_LIKE_ROLES and procedure_type in ('upper_gi', 'colonoscopy'):
            sched_cap = s['scheduler_upper_gi_cap'] if procedure_type == 'upper_gi' else s['scheduler_colono_cap']
            sched_count = c['scheduler_upper_gi'] if procedure_type == 'upper_gi' else c['scheduler_colono']
            if sched_count >= sched_cap:
                proc_label = PROCEDURE_LABELS[procedure_type]
                return False, (
                    f'Scheduler/Endoscopy Staff quota reached for {proc_label} on {d_iso} '
                    f'({sched_count}/{sched_cap}). Remaining slots are reserved for on-call '
                    f'doctors and inpatients.'
                ), {}

        return True, None, {}

    return False, 'Unhandled procedure type.', {}


# ----------------------------------------------------------------------
# Auth routes
# ----------------------------------------------------------------------
@app.route('/')
def index():
    if current_user():
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/register', methods=['GET', 'POST'])
def register():
    dbconn = get_db()
    if request.method == 'POST':
        username = request.form.get('username', '').strip().lower()
        full_name = request.form.get('full_name', '').strip()
        password = request.form.get('password', '')
        role = request.form.get('role', '')

        if role not in ALL_ROLES:
            flash('Please choose a valid role.', 'error')
            return redirect(url_for('register'))
        if not username or not full_name or len(password) < 6:
            flash('All fields are required and password must be at least 6 characters.', 'error')
            return redirect(url_for('register'))
        existing = dbconn.execute('SELECT id FROM user WHERE username = ?', (username,)).fetchone()
        if existing:
            flash('That username is already taken.', 'error')
            return redirect(url_for('register'))
        if role == ROLE_ONCALL:
            count = dbconn.execute('SELECT COUNT(*) AS c FROM user WHERE role = ?', (ROLE_ONCALL,)).fetchone()['c']
            if count >= ONCALL_SLOT_COUNT:
                flash(f'All {ONCALL_SLOT_COUNT} on-call doctor accounts already exist. Contact Admin.', 'error')
                return redirect(url_for('register'))

        dbconn.execute(
            'INSERT INTO user (username, full_name, password_hash, role, is_approved, created_at) '
            'VALUES (?, ?, ?, ?, 0, ?)',
            (username, full_name, generate_password_hash(password), role, datetime.utcnow().isoformat())
        )
        dbconn.commit()
        flash('Account created. It is now pending approval by an Admin or Specialist before you can log in.', 'success')
        return redirect(url_for('login'))

    return render_template('register.html', roles=ALL_ROLES, role_labels=ROLE_LABELS)


@app.route('/login', methods=['GET', 'POST'])
def login():
    next_url = request.values.get('next', '')
    if request.method == 'POST':
        username = request.form.get('username', '').strip().lower()
        password = request.form.get('password', '')
        row = get_db().execute('SELECT * FROM user WHERE username = ?', (username,)).fetchone()
        if not row or not check_password_hash(row['password_hash'], password):
            flash('Invalid username or password.', 'error')
            return redirect(url_for('login', next=next_url))
        if not row['is_approved']:
            flash('Your account is pending approval by an Admin or Specialist.', 'error')
            return redirect(url_for('login', next=next_url))
        session.permanent = False  # session cookie only - cleared when the browser is closed
        session['user_id'] = row['id']
        if next_url and next_url.startswith('/') and not next_url.startswith('//'):
            return redirect(next_url)
        return redirect(url_for('dashboard'))
    return render_template('login.html', next=next_url)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        user = current_user()
        current_pw = request.form.get('current_password', '')
        new_pw = request.form.get('new_password', '')
        confirm_pw = request.form.get('confirm_password', '')

        if not check_password_hash(user['password_hash'], current_pw):
            flash('Current password is incorrect.', 'error')
            return redirect(url_for('change_password'))
        if len(new_pw) < 6:
            flash('New password must be at least 6 characters.', 'error')
            return redirect(url_for('change_password'))
        if new_pw != confirm_pw:
            flash('New password and confirmation do not match.', 'error')
            return redirect(url_for('change_password'))

        dbconn = get_db()
        dbconn.execute('UPDATE user SET password_hash = ? WHERE id = ?',
                        (generate_password_hash(new_pw), user['id']))
        dbconn.commit()
        flash('Password updated successfully.', 'success')
        return redirect(url_for('dashboard'))

    return render_template('change_password.html')


# ----------------------------------------------------------------------
# Core app routes
# ----------------------------------------------------------------------
@app.route('/dashboard')
@login_required
def dashboard():
    user = current_user()
    today = today_pk()
    return render_template(
        'dashboard.html',
        today=today.isoformat(),
        today_display=today.strftime('%A, %d %B %Y'),
        can_override=user['role'] in CAN_OVERRIDE,
        can_book_ercp=user['role'] in CAN_BOOK_ERCP,
        can_book_special=user['role'] in CAN_BOOK_SPECIAL,
        is_nurse_manager=user['role'] == ROLE_NURSE_MANAGER,
        is_ercp_day=today.weekday() in ERCP_WEEKDAYS,
    )


@app.route('/calendar')
@login_required
def calendar_view():
    user = current_user()
    today = today_pk()
    return render_template(
        'calendar.html',
        today=today.isoformat(),
        can_override=user['role'] in CAN_OVERRIDE,
        can_book_ercp=user['role'] in CAN_BOOK_ERCP,
        can_book_special=user['role'] in CAN_BOOK_SPECIAL,
    )


@app.route('/admin')
@roles_required(*CAN_APPROVE)
def admin_panel():
    dbconn = get_db()
    pending = dbconn.execute('SELECT * FROM user WHERE is_approved = 0 ORDER BY created_at').fetchall()
    approved = dbconn.execute('SELECT * FROM user WHERE is_approved = 1 ORDER BY role, username').fetchall()
    settings_row = get_settings()
    return render_template('admin.html', pending=pending, approved=approved, settings=settings_row)


@app.route('/admin/approve/<int:user_id>', methods=['POST'])
@roles_required(*CAN_APPROVE)
def approve_user(user_id):
    dbconn = get_db()
    row = dbconn.execute('SELECT * FROM user WHERE id = ?', (user_id,)).fetchone()
    if not row:
        flash('User not found.', 'error')
        return redirect(url_for('admin_panel'))
    dbconn.execute('UPDATE user SET is_approved = 1 WHERE id = ?', (user_id,))
    dbconn.commit()
    flash(f'Approved account "{row["username"]}" ({user_role_label(row["role"])}).', 'success')
    return redirect(url_for('admin_panel'))


@app.route('/admin/reject/<int:user_id>', methods=['POST'])
@roles_required(*CAN_APPROVE)
def reject_user(user_id):
    dbconn = get_db()
    dbconn.execute('DELETE FROM user WHERE id = ?', (user_id,))
    dbconn.commit()
    flash('Pending registration rejected and removed.', 'success')
    return redirect(url_for('admin_panel'))


@app.route('/admin/delete-user/<int:user_id>', methods=['POST'])
@roles_required(*CAN_APPROVE)
def delete_user(user_id):
    dbconn = get_db()
    row = dbconn.execute('SELECT * FROM user WHERE id = ?', (user_id,)).fetchone()
    if not row:
        flash('User not found.', 'error')
        return redirect(url_for('admin_panel'))
    if row['role'] == ROLE_ADMIN:
        other_admins = dbconn.execute(
            "SELECT COUNT(*) AS c FROM user WHERE role = 'admin' AND id != ?", (user_id,)
        ).fetchone()['c']
        if other_admins == 0:
            flash('Cannot delete the last remaining Admin account.', 'error')
            return redirect(url_for('admin_panel'))
    dbconn.execute('DELETE FROM user WHERE id = ?', (user_id,))
    dbconn.commit()
    flash(f'Deleted account "{row["username"]}" ({user_role_label(row["role"])}).', 'success')
    if row['id'] == session.get('user_id'):
        session.clear()
        return redirect(url_for('login'))
    return redirect(url_for('admin_panel'))


@app.route('/admin/settings', methods=['POST'])
@roles_required(ROLE_ADMIN)
def update_settings():
    dbconn = get_db()
    try:
        vals = (
            int(request.form.get('global_upper_gi_cap')),
            int(request.form.get('global_colono_cap')),
            int(request.form.get('global_peg_cap')),
            int(request.form.get('scheduler_upper_gi_cap')),
            int(request.form.get('scheduler_colono_cap')),
            int(request.form.get('warning_threshold')),
        )
        dbconn.execute(
            'UPDATE settings SET global_upper_gi_cap=?, global_colono_cap=?, global_peg_cap=?, '
            'scheduler_upper_gi_cap=?, scheduler_colono_cap=?, warning_threshold=? WHERE id = 1',
            vals
        )
        dbconn.commit()
        flash('System caps updated.', 'success')
    except (ValueError, TypeError):
        flash('Caps must be valid integers.', 'error')
    return redirect(url_for('admin_panel'))


@app.route('/print/<date_str>')
@login_required
def print_list(date_str):
    try:
        d = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return redirect(url_for('dashboard'))

    only = request.args.get('only', '')  # '', 'ercp', or 'other'
    if only == 'ercp':
        rows = get_db().execute(
            'SELECT * FROM appointment WHERE appointment_date = ? AND procedure_type = ? '
            'ORDER BY procedure_type', (d.isoformat(), 'ercp')
        ).fetchall()
        list_title = 'ERCP List'
    elif only == 'other':
        rows = get_db().execute(
            'SELECT * FROM appointment WHERE appointment_date = ? AND procedure_type != ? '
            'ORDER BY procedure_type', (d.isoformat(), 'ercp')
        ).fetchall()
        list_title = 'Endoscopy Suite — Daily Case List'
    else:
        rows = get_db().execute(
            'SELECT * FROM appointment WHERE appointment_date = ? ORDER BY procedure_type', (d.isoformat(),)
        ).fetchall()
        list_title = 'Endoscopy Suite — Daily Case List'

    return render_template('print_list.html', appts=rows, day=d, list_title=list_title)


@app.route('/statistics')
@roles_required(*CAN_VIEW_STATS)
def statistics_view():
    start_str = request.args.get('start', '')
    end_str = request.args.get('end', '')
    stats = None
    error = None

    if start_str and end_str:
        try:
            start_d = datetime.strptime(start_str, '%Y-%m-%d').date()
            end_d = datetime.strptime(end_str, '%Y-%m-%d').date()
            if end_d < start_d:
                error = 'End date must be on or after the start date.'
            else:
                stats = compute_statistics(start_d, end_d)
        except ValueError:
            error = 'Please provide valid start and end dates.'

    return render_template(
        'statistics.html', stats=stats, start=start_str, end=end_str, error=error
    )


@app.route('/admin/holidays')
@roles_required(*CAN_MANAGE_HOLIDAYS)
def holidays_view():
    dbconn = get_db()
    today_iso = today_pk().isoformat()
    upcoming = dbconn.execute(
        'SELECT * FROM holiday WHERE holiday_date >= ? ORDER BY holiday_date', (today_iso,)
    ).fetchall()
    past = dbconn.execute(
        'SELECT * FROM holiday WHERE holiday_date < ? ORDER BY holiday_date DESC LIMIT 20', (today_iso,)
    ).fetchall()
    return render_template('holidays.html', upcoming=upcoming, past=past)


@app.route('/admin/holidays/sync', methods=['POST'])
@roles_required(*CAN_MANAGE_HOLIDAYS)
def sync_holidays_route():
    try:
        events = fetch_and_parse_holidays()
        count = sync_holidays_into_db(get_db(), events)
        flash(f'Synced {count} public holidays for Pakistan (via Calendarific).', 'success')
    except Exception as e:
        flash(
            f'Holiday sync failed: {e}. If your server\'s network cannot reach external '
            f'APIs (common on some free hosting tiers), add holidays manually below instead.',
            'error'
        )
    return redirect(url_for('holidays_view'))


@app.route('/admin/holidays/add', methods=['POST'])
@roles_required(*CAN_MANAGE_HOLIDAYS)
def add_holiday_route():
    date_str = request.form.get('holiday_date', '')
    name = request.form.get('name', '').strip()
    try:
        d = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Please provide a valid date.', 'error')
        return redirect(url_for('holidays_view'))
    if not name:
        flash('Please provide a holiday name.', 'error')
        return redirect(url_for('holidays_view'))

    dbconn = get_db()
    dbconn.execute(
        'INSERT INTO holiday (holiday_date, name, source, created_at) VALUES (?,?,?,?) '
        'ON CONFLICT(holiday_date) DO UPDATE SET name=excluded.name, source=excluded.source',
        (d.isoformat(), name, 'manual', datetime.utcnow().isoformat())
    )
    dbconn.commit()
    flash(f'Added holiday: {name} on {d.isoformat()}.', 'success')
    return redirect(url_for('holidays_view'))


@app.route('/admin/holidays/delete/<int:holiday_id>', methods=['POST'])
@roles_required(*CAN_MANAGE_HOLIDAYS)
def delete_holiday_route(holiday_id):
    dbconn = get_db()
    dbconn.execute('DELETE FROM holiday WHERE id = ?', (holiday_id,))
    dbconn.commit()
    flash('Holiday removed.', 'success')
    return redirect(url_for('holidays_view'))


@app.route('/admin/export')
@roles_required(*CAN_OVERRIDE)
def export_appointments():
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    start_str = request.args.get('start', '')
    end_str = request.args.get('end', '')
    try:
        start_d = datetime.strptime(start_str, '%Y-%m-%d').date()
        end_d = datetime.strptime(end_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Please provide a valid start and end date to export.', 'error')
        return redirect(url_for('admin_panel'))
    if end_d < start_d:
        flash('End date must be on or after the start date.', 'error')
        return redirect(url_for('admin_panel'))

    rows = get_db().execute(
        'SELECT * FROM appointment WHERE appointment_date BETWEEN ? AND ? '
        'ORDER BY appointment_date, procedure_type',
        (start_d.isoformat(), end_d.isoformat())
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Bookings'

    headers = ['Date', 'Procedure', 'Patient Name', 'Gender', 'Age', 'Phone', 'MRN', 'Referral',
               'Clinical Notes', 'Hb (on admission)', 'Platelet', 'INR', 'Comorbidities / Etiology',
               'Bleeding', 'Override', 'Booked By', 'Role', 'Created At']
    ws.append(headers)
    header_fill = PatternFill('solid', start_color='A6192E', end_color='A6192E')
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for r in rows:
        ws.append([
            r['appointment_date'],
            PROCEDURE_LABELS.get(r['procedure_type'], r['procedure_type']),
            r['patient_name'],
            r['gender'],
            r['age'],
            r['phone'],
            r['mrn'],
            r['referral'],
            r['clinical_notes'],
            r['on_admission_hb'],
            r['platelet'],
            r['inr'],
            r['comorbs_etiology'],
            'Yes' if r['is_bleeding'] else 'No',
            'Yes' if r['is_override'] else 'No',
            r['booked_by_username'],
            user_role_label(r['booked_by_role']),
            r['created_at'],
        ])

    widths = [12, 20, 22, 10, 6, 15, 14, 18, 34, 12, 10, 8, 24, 10, 10, 16, 24, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{max(ws.max_row, 1)}'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import send_file
    filename = f'gastro_bookings_{start_d.isoformat()}_to_{end_d.isoformat()}.xlsx'
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# ----------------------------------------------------------------------
# JSON API
# ----------------------------------------------------------------------
@app.route('/api/me')
@login_required
def api_me():
    u = current_user()
    return jsonify({
        'username': u['username'],
        'full_name': u['full_name'],
        'role': u['role'],
        'role_label': user_role_label(u['role']),
        'can_override': u['role'] in CAN_OVERRIDE,
        'can_book_ercp': u['role'] in CAN_BOOK_ERCP,
        'can_book_special': u['role'] in CAN_BOOK_SPECIAL,
    })


@app.route('/api/day/<date_str>')
@login_required
def api_day(date_str):
    try:
        d = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'invalid date'}), 400
    d_iso = d.isoformat()
    rows = get_db().execute(
        'SELECT * FROM appointment WHERE appointment_date = ? ORDER BY created_at', (d_iso,)
    ).fetchall()
    info = day_status(d_iso)
    return jsonify({
        'date': d_iso,
        'is_ercp_day': d.weekday() in ERCP_WEEKDAYS,
        'status': info['status'],
        'counts': info['counts'],
        'caps': info['caps'],
        'is_sunday': info['is_sunday'],
        'holiday_name': info['holiday_name'],
        'warning_threshold': get_settings()['warning_threshold'],
        'appointments': [appt_to_dict(r) for r in rows],
    })


@app.route('/api/month-summary')
@login_required
def api_month_summary():
    year = int(request.args.get('year'))
    month = int(request.args.get('month'))
    first_day = date(year, month, 1)
    if month == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month + 1, 1)
    days = {}
    d = first_day
    while d < next_month:
        info = day_status(d.isoformat())
        days[d.isoformat()] = {
            'status': info['status'],
            'upper_gi': info['counts']['upper_gi'],
            'colonoscopy': info['counts']['colonoscopy'],
            'peg_tube': info['counts']['peg_tube'],
            'ercp': info['counts']['ercp'],
            'is_ercp_day': d.weekday() in ERCP_WEEKDAYS,
            'is_sunday': info['is_sunday'],
            'holiday_name': info['holiday_name'],
        }
        d += timedelta(days=1)
    return jsonify(days)


@app.route('/api/book', methods=['POST'])
@login_required
def api_book():
    user = current_user()
    payload = request.get_json(force=True, silent=True) or {}

    ok, error, extra = validate_booking(user, payload)
    if not ok:
        return jsonify({'error': error}), 400

    d = datetime.strptime(payload['appointment_date'], '%Y-%m-%d').date()
    dbconn = get_db()
    cur = dbconn.execute(
        'INSERT INTO appointment '
        '(patient_name, gender, age, phone, mrn, clinical_notes, on_admission_hb, platelet, inr, '
        'comorbs_etiology, referral, procedure_type, appointment_date, is_bleeding, is_override, '
        'booked_by_username, booked_by_role, created_at) '
        'VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',
        (
            payload['patient_name'].strip(),
            payload['gender'],
            int(payload['age']),
            payload['phone'].strip(),
            (payload.get('mrn') or '').strip(),
            (payload.get('clinical_notes') or '').strip(),
            (payload.get('on_admission_hb') or '').strip(),
            (payload.get('platelet') or '').strip(),
            (payload.get('inr') or '').strip(),
            (payload.get('comorbs_etiology') or '').strip(),
            (payload.get('referral') or '').strip(),
            payload['procedure_type'],
            d.isoformat(),
            1 if payload.get('is_bleeding') else 0,
            1 if extra.get('override_used') else 0,
            user['username'],
            user['role'],
            datetime.utcnow().isoformat(),
        )
    )
    dbconn.commit()
    new_row = dbconn.execute('SELECT * FROM appointment WHERE id = ?', (cur.lastrowid,)).fetchone()

    response = {'success': True, 'appointment': appt_to_dict(new_row)}
    if extra.get('warning'):
        response['warning'] = (
            f"Regular case list for {d.isoformat()} now exceeds "
            f"{get_settings()['warning_threshold']} cases."
        )
    return jsonify(response)


@app.route('/api/appointment/<int:appt_id>', methods=['PUT'])
@login_required
def api_edit_appointment(appt_id):
    user = current_user()
    dbconn = get_db()
    row = dbconn.execute('SELECT * FROM appointment WHERE id = ?', (appt_id,)).fetchone()
    if not row:
        return jsonify({'error': 'Not found.'}), 404
    if row['booked_by_username'] != user['username'] and user['role'] not in CAN_OVERRIDE:
        return jsonify({'error': 'You do not have permission to edit this booking.'}), 403
    if user['role'] in SCHEDULER_LIKE_ROLES and is_time_locked(row):
        return jsonify({'error': (
            'Scheduler/Endoscopy Staff accounts can no longer edit this booking — the appointment day has '
            'arrived or 48 hours have passed since it was booked. Contact an Admin or Specialist.'
        )}), 403

    payload = request.get_json(force=True, silent=True) or {}
    ok, error, extra = validate_booking(user, payload, exclude_id=appt_id)
    if not ok:
        return jsonify({'error': error}), 400

    d = datetime.strptime(payload['appointment_date'], '%Y-%m-%d').date()
    dbconn.execute(
        'UPDATE appointment SET patient_name=?, gender=?, age=?, phone=?, mrn=?, '
        'clinical_notes=?, on_admission_hb=?, platelet=?, inr=?, comorbs_etiology=?, referral=?, '
        'procedure_type=?, appointment_date=?, is_bleeding=?, is_override=? '
        'WHERE id = ?',
        (
            payload['patient_name'].strip(),
            payload['gender'],
            int(payload['age']),
            payload['phone'].strip(),
            (payload.get('mrn') or '').strip(),
            (payload.get('clinical_notes') or '').strip(),
            (payload.get('on_admission_hb') or '').strip(),
            (payload.get('platelet') or '').strip(),
            (payload.get('inr') or '').strip(),
            (payload.get('comorbs_etiology') or '').strip(),
            (payload.get('referral') or '').strip(),
            payload['procedure_type'],
            d.isoformat(),
            1 if payload.get('is_bleeding') else 0,
            1 if extra.get('override_used') else 0,
            appt_id,
        )
    )
    dbconn.commit()
    updated = dbconn.execute('SELECT * FROM appointment WHERE id = ?', (appt_id,)).fetchone()

    response = {'success': True, 'appointment': appt_to_dict(updated)}
    if extra.get('warning'):
        response['warning'] = (
            f"Regular case list for {d.isoformat()} now exceeds "
            f"{get_settings()['warning_threshold']} cases."
        )
    return jsonify(response)


@app.route('/api/appointment/<int:appt_id>', methods=['DELETE'])
@login_required
def api_delete_appointment(appt_id):
    user = current_user()
    dbconn = get_db()
    row = dbconn.execute('SELECT * FROM appointment WHERE id = ?', (appt_id,)).fetchone()
    if not row:
        return jsonify({'error': 'Not found.'}), 404
    if row['booked_by_username'] != user['username'] and user['role'] not in CAN_OVERRIDE:
        return jsonify({'error': 'You do not have permission to cancel this booking.'}), 403
    if is_time_locked(row) and user['role'] != ROLE_ADMIN:
        return jsonify({'error': (
            'This booking can only be deleted by an Admin once the appointment day has '
            'arrived or 48 hours have passed since it was booked.'
        )}), 403
    dbconn.execute('DELETE FROM appointment WHERE id = ?', (appt_id,))
    dbconn.commit()
    return jsonify({'success': True})


@app.route('/api/appointment/<int:appt_id>/no-show', methods=['POST'])
@login_required
def api_toggle_no_show(appt_id):
    """Toggle the 'Did Not Show Up' flag. Open to every role — no ownership
    or time restriction, per department policy."""
    dbconn = get_db()
    row = dbconn.execute('SELECT * FROM appointment WHERE id = ?', (appt_id,)).fetchone()
    if not row:
        return jsonify({'error': 'Not found.'}), 404
    new_value = 0 if row['no_show'] else 1
    dbconn.execute('UPDATE appointment SET no_show = ? WHERE id = ?', (new_value, appt_id))
    dbconn.commit()
    updated = dbconn.execute('SELECT * FROM appointment WHERE id = ?', (appt_id,)).fetchone()
    return jsonify({'success': True, 'appointment': appt_to_dict(updated)})


# ----------------------------------------------------------------------
# ERCP Reporting Module
# ----------------------------------------------------------------------
def ercp_report_to_dict(row):
    return dict(row)


@app.route('/ercp/<int:appointment_id>')
@roles_required(*CAN_ACCESS_ERCP_REPORTS)
def ercp_report_view(appointment_id):
    dbconn = get_db()
    appt = dbconn.execute('SELECT * FROM appointment WHERE id = ?', (appointment_id,)).fetchone()
    if not appt or appt['procedure_type'] != 'ercp':
        flash('That is not an ERCP appointment.', 'error')
        return redirect(url_for('dashboard'))

    user = current_user()
    report = get_or_create_ercp_report(appointment_id, user['username'])
    research = dbconn.execute('SELECT * FROM ercp_research WHERE report_id = ?', (report['id'],)).fetchone()
    images = dbconn.execute(
        'SELECT * FROM ercp_report_image WHERE report_id = ? ORDER BY slot', (report['id'],)
    ).fetchall()
    image_by_slot = {img['slot']: img for img in images}
    endoscopists = dbconn.execute(
        'SELECT * FROM endoscopist WHERE is_active = 1 OR id = ? ORDER BY full_name',
        (report['endoscopist_id'] or 0,)
    ).fetchall()
    anesthesiologists = [
        row['anesthesiologist'] for row in dbconn.execute(
            "SELECT DISTINCT anesthesiologist FROM ercp_report "
            "WHERE anesthesiologist != '' ORDER BY anesthesiologist"
        ).fetchall()
    ]

    return render_template(
        'ercp_report.html',
        appt=appt, report=report, research=research,
        image_by_slot=image_by_slot, image_slots=range(1, ERCP_IMAGE_SLOTS + 1),
        endoscopists=endoscopists, anesthesiologists=anesthesiologists,
        sedation_options=SEDATION_OPTIONS, indication_options=INDICATION_OPTIONS,
        duodenoscope_advancement_options=DUODENOSCOPE_ADVANCEMENT_OPTIONS,
        papilla_shape_options=PAPILLA_SHAPE_OPTIONS,
        papilla_location_options=PAPILLA_LOCATION_OPTIONS,
        papilla_access_options=PAPILLA_ACCESS_OPTIONS,
        cannulation_options=CANNULATION_OPTIONS,
        cholangiogram_categories=CHOLANGIOGRAM_CATEGORIES, therapeutic_options=THERAPEUTIC_OPTIONS,
        biopsy_options=BIOPSY_OPTIONS, complication_options=COMPLICATION_OPTIONS,
        stent_type_options=STENT_TYPE_OPTIONS, stent_diameter_options=STENT_DIAMETER_OPTIONS,
        stent_length_options=STENT_LENGTH_OPTIONS, stent_location_options=STENT_LOCATION_OPTIONS,
        stent_deployment_options=STENT_DEPLOYMENT_OPTIONS, stent_drainage_options=STENT_DRAINAGE_OPTIONS,
        is_locked=(report['status'] == 'finalized'),
        can_unlock=(user['role'] == ROLE_ADMIN),
    )


@app.route('/ercp/<int:report_id>/save', methods=['POST'])
@roles_required(*CAN_ACCESS_ERCP_REPORTS)
def ercp_report_save(report_id):
    dbconn = get_db()
    report = dbconn.execute('SELECT * FROM ercp_report WHERE id = ?', (report_id,)).fetchone()
    if not report:
        return jsonify({'error': 'Report not found.'}), 404
    if report['status'] == 'finalized':
        return jsonify({'error': 'This report is finalized and read-only. Ask an Admin to unlock it first.'}), 403

    payload = request.get_json(force=True, silent=True) or {}

    def multi(key):
        vals = payload.get(key) or []
        if isinstance(vals, str):
            vals = [vals]
        return ', '.join(v.strip() for v in vals if v and v.strip())

    fields = {
        'endoscopist_id': payload.get('endoscopist_id') or None,
        'assistants': (payload.get('assistants') or '').strip(),
        'technician': (payload.get('technician') or '').strip(),
        'sedation': payload.get('sedation') or '',
        'anesthesiologist': (payload.get('anesthesiologist') or '').strip(),
        'indication': payload.get('indication') or '',
        'duodenoscope_advancement': payload.get('duodenoscope_advancement') or '',
        'papilla': payload.get('papilla') or '',
        'papilla_location': payload.get('papilla_location') or '',
        'papilla_access': payload.get('papilla_access') or '',
        'cannulation': payload.get('cannulation') or '',
        'cholangiogram_findings': multi('cholangiogram_findings'),
        'cholangio_cbd_mm': (payload.get('cholangio_cbd_mm') or '').strip(),
        'cholangio_chd_mm': (payload.get('cholangio_chd_mm') or '').strip(),
        'cholangio_rhd_mm': (payload.get('cholangio_rhd_mm') or '').strip(),
        'cholangio_lhd_mm': (payload.get('cholangio_lhd_mm') or '').strip(),
        'cholangio_largest_stone_mm': (payload.get('cholangio_largest_stone_mm') or '').strip(),
        'cholangio_stone_count': (payload.get('cholangio_stone_count') or '').strip(),
        'cholangio_stricture_length_mm': (payload.get('cholangio_stricture_length_mm') or '').strip(),
        'therapeutic_procedures': multi('therapeutic_procedures'),
        'stent_placed': payload.get('stent_placed') or '',
        'stent_type': payload.get('stent_type') or '',
        'stent_manufacturer': (payload.get('stent_manufacturer') or '').strip(),
        'stent_diameter': payload.get('stent_diameter') or '',
        'stent_length': payload.get('stent_length') or '',
        'stent_count': (payload.get('stent_count') or '').strip(),
        'stent_location': payload.get('stent_location') or '',
        'stent_deployment': payload.get('stent_deployment') or '',
        'stent_drainage': payload.get('stent_drainage') or '',
        'biopsy': payload.get('biopsy') or '',
        'complications': multi('complications'),
        'procedure_note': payload.get('procedure_note') or '',
        'impression': (payload.get('impression') or '').strip(),
        'recommendations': (payload.get('recommendations') or '').strip(),
        'lab_total_bilirubin': (payload.get('lab_total_bilirubin') or '').strip(),
        'lab_direct_bilirubin': (payload.get('lab_direct_bilirubin') or '').strip(),
        'lab_alt': (payload.get('lab_alt') or '').strip(),
        'lab_ast': (payload.get('lab_ast') or '').strip(),
        'lab_alp': (payload.get('lab_alp') or '').strip(),
        'lab_ggt': (payload.get('lab_ggt') or '').strip(),
        'lab_albumin': (payload.get('lab_albumin') or '').strip(),
        'lab_hb': (payload.get('lab_hb') or '').strip(),
        'lab_wbc': (payload.get('lab_wbc') or '').strip(),
        'lab_platelets': (payload.get('lab_platelets') or '').strip(),
        'lab_pt': (payload.get('lab_pt') or '').strip(),
        'lab_inr': (payload.get('lab_inr') or '').strip(),
        'lab_creatinine': (payload.get('lab_creatinine') or '').strip(),
        'imaging_us': (payload.get('imaging_us') or '').strip(),
        'imaging_ct': (payload.get('imaging_ct') or '').strip(),
        'imaging_mrcp': (payload.get('imaging_mrcp') or '').strip(),
    }
    set_clause = ', '.join(f'{k}=?' for k in fields)
    dbconn.execute(
        f'UPDATE ercp_report SET {set_clause}, updated_at=? WHERE id=?',
        (*fields.values(), datetime.utcnow().isoformat(), report_id)
    )

    research_payload = payload.get('research') or {}
    research_fields = {
        'fluoro_time_sec': research_payload.get('fluoro_time_sec', ''),
        'contrast_volume_ml': research_payload.get('contrast_volume_ml', ''),
        'cbd_diameter_mm': research_payload.get('cbd_diameter_mm', ''),
        'stone_size_mm': research_payload.get('stone_size_mm', ''),
        'stone_count': research_payload.get('stone_count', ''),
        'stone_clearance': research_payload.get('stone_clearance', ''),
        'pd_findings': research_payload.get('pd_findings', ''),
        'pd_intervention': research_payload.get('pd_intervention', ''),
        'device_details': research_payload.get('device_details', ''),
        'procedure_duration_min': research_payload.get('procedure_duration_min', ''),
        'asa_class': research_payload.get('asa_class', ''),
        'complication_severity': research_payload.get('complication_severity', ''),
        'disposition': research_payload.get('disposition', ''),
        'followup_plan': research_payload.get('followup_plan', ''),
    }
    r_set_clause = ', '.join(f'{k}=?' for k in research_fields)
    dbconn.execute(
        f'UPDATE ercp_research SET {r_set_clause}, updated_at=? WHERE report_id=?',
        (*research_fields.values(), datetime.utcnow().isoformat(), report_id)
    )
    dbconn.commit()
    return jsonify({'success': True})


@app.route('/ercp/<int:report_id>/generate-note', methods=['POST'])
@roles_required(*CAN_ACCESS_ERCP_REPORTS)
def ercp_generate_note(report_id):
    payload = request.get_json(force=True, silent=True) or {}

    def multi(key):
        vals = payload.get(key) or []
        if isinstance(vals, str):
            vals = [vals]
        return ', '.join(v.strip() for v in vals if v and v.strip())

    note = generate_procedure_note({
        'sedation': payload.get('sedation') or '',
        'indication': payload.get('indication') or '',
        'duodenoscope_advancement': payload.get('duodenoscope_advancement') or '',
        'papilla': payload.get('papilla') or '',
        'papilla_location': payload.get('papilla_location') or '',
        'papilla_access': payload.get('papilla_access') or '',
        'cannulation': payload.get('cannulation') or '',
        'cholangiogram_findings': multi('cholangiogram_findings'),
        'cholangio_cbd_mm': payload.get('cholangio_cbd_mm') or '',
        'cholangio_chd_mm': payload.get('cholangio_chd_mm') or '',
        'cholangio_rhd_mm': payload.get('cholangio_rhd_mm') or '',
        'cholangio_lhd_mm': payload.get('cholangio_lhd_mm') or '',
        'cholangio_largest_stone_mm': payload.get('cholangio_largest_stone_mm') or '',
        'cholangio_stone_count': payload.get('cholangio_stone_count') or '',
        'cholangio_stricture_length_mm': payload.get('cholangio_stricture_length_mm') or '',
        'therapeutic_procedures': multi('therapeutic_procedures'),
        'stent_placed': payload.get('stent_placed') or '',
        'stent_type': payload.get('stent_type') or '',
        'stent_manufacturer': payload.get('stent_manufacturer') or '',
        'stent_diameter': payload.get('stent_diameter') or '',
        'stent_length': payload.get('stent_length') or '',
        'stent_count': payload.get('stent_count') or '',
        'stent_location': payload.get('stent_location') or '',
        'stent_deployment': payload.get('stent_deployment') or '',
        'stent_drainage': payload.get('stent_drainage') or '',
        'biopsy': payload.get('biopsy') or '',
        'complications': multi('complications'),
    })
    return jsonify({'note': note})


@app.route('/ercp/<int:report_id>/finalize', methods=['POST'])
@roles_required(*CAN_ACCESS_ERCP_REPORTS)
def ercp_finalize(report_id):
    dbconn = get_db()
    report = dbconn.execute('SELECT * FROM ercp_report WHERE id = ?', (report_id,)).fetchone()
    if not report:
        return jsonify({'error': 'Report not found.'}), 404
    if report['status'] == 'finalized':
        return jsonify({'error': 'Already finalized.'}), 400
    if not report['endoscopist_id']:
        return jsonify({'error': 'Please select the Endoscopist before finalizing.'}), 400

    user = current_user()
    now = datetime.utcnow().isoformat()
    dbconn.execute(
        "UPDATE ercp_report SET status='finalized', finalized_by=?, finalized_at=?, updated_at=? WHERE id=?",
        (user['username'], now, now, report_id)
    )
    dbconn.commit()
    return jsonify({'success': True})


@app.route('/ercp/<int:report_id>/unlock', methods=['POST'])
@roles_required(ROLE_ADMIN)
def ercp_unlock(report_id):
    dbconn = get_db()
    report = dbconn.execute('SELECT * FROM ercp_report WHERE id = ?', (report_id,)).fetchone()
    if not report:
        flash('Report not found.', 'error')
        return redirect(url_for('dashboard'))
    user = current_user()
    now = datetime.utcnow().isoformat()
    dbconn.execute(
        "UPDATE ercp_report SET status='draft', unlocked_by=?, unlocked_at=?, updated_at=? WHERE id=?",
        (user['username'], now, now, report_id)
    )
    dbconn.commit()
    flash('Report unlocked for editing.', 'success')
    return redirect(url_for('ercp_report_view', appointment_id=report['appointment_id']))


@app.route('/ercp/<int:report_id>/image/<int:slot>', methods=['POST'])
@roles_required(*CAN_ACCESS_ERCP_REPORTS)
def ercp_upload_image(report_id, slot):
    dbconn = get_db()
    report = dbconn.execute('SELECT * FROM ercp_report WHERE id = ?', (report_id,)).fetchone()
    if not report:
        return jsonify({'error': 'Report not found.'}), 404
    if report['status'] == 'finalized':
        return jsonify({'error': 'This report is finalized and read-only.'}), 403
    if slot < 1 or slot > ERCP_IMAGE_SLOTS:
        return jsonify({'error': 'Invalid image slot.'}), 400
    file = request.files.get('image')
    if not file or not file.filename:
        return jsonify({'error': 'No image file provided.'}), 400

    user = current_user()
    filename = f'report_{report_id}_slot_{slot}.jpg'
    dest_path = os.path.join(ERCP_IMAGES_DIR, filename)
    try:
        compress_ercp_image(file, dest_path)
    except ModuleNotFoundError:
        return jsonify({'error': (
            'Image support (Pillow) is not installed on the server yet. Run '
            '"pip install -r requirements.txt" inside the project\'s virtualenv on '
            'PythonAnywhere, then reload the web app, and try uploading again.'
        )}), 500
    except Exception as e:
        return jsonify({'error': f'Could not process image: {e}'}), 400

    now = datetime.utcnow().isoformat()
    dbconn.execute(
        'INSERT INTO ercp_report_image (report_id, slot, filename, uploaded_by, uploaded_at) '
        'VALUES (?,?,?,?,?) '
        'ON CONFLICT(report_id, slot) DO UPDATE SET filename=excluded.filename, '
        'uploaded_by=excluded.uploaded_by, uploaded_at=excluded.uploaded_at',
        (report_id, slot, filename, user['username'], now)
    )
    dbconn.commit()
    return jsonify({'success': True, 'slot': slot})


@app.route('/ercp/<int:report_id>/image/<int:slot>', methods=['GET'])
@roles_required(*CAN_ACCESS_ERCP_REPORTS)
def ercp_view_image(report_id, slot):
    from flask import send_from_directory
    dbconn = get_db()
    img = dbconn.execute(
        'SELECT * FROM ercp_report_image WHERE report_id = ? AND slot = ?', (report_id, slot)
    ).fetchone()
    if not img:
        return jsonify({'error': 'Not found.'}), 404
    return send_from_directory(ERCP_IMAGES_DIR, img['filename'])


@app.route('/ercp/<int:report_id>/image/<int:slot>/delete', methods=['POST'])
@roles_required(*CAN_ACCESS_ERCP_REPORTS)
def ercp_delete_image(report_id, slot):
    dbconn = get_db()
    report = dbconn.execute('SELECT * FROM ercp_report WHERE id = ?', (report_id,)).fetchone()
    if not report:
        return jsonify({'error': 'Report not found.'}), 404
    if report['status'] == 'finalized':
        return jsonify({'error': 'This report is finalized and read-only.'}), 403
    img = dbconn.execute(
        'SELECT * FROM ercp_report_image WHERE report_id = ? AND slot = ?', (report_id, slot)
    ).fetchone()
    if img:
        try:
            os.remove(os.path.join(ERCP_IMAGES_DIR, img['filename']))
        except OSError:
            pass
        dbconn.execute('DELETE FROM ercp_report_image WHERE id = ?', (img['id'],))
        dbconn.commit()
    return jsonify({'success': True})


@app.route('/ercp/<int:report_id>/print')
@roles_required(*CAN_ACCESS_ERCP_REPORTS)
def ercp_print(report_id):
    dbconn = get_db()
    report = dbconn.execute('SELECT * FROM ercp_report WHERE id = ?', (report_id,)).fetchone()
    if not report:
        flash('Report not found.', 'error')
        return redirect(url_for('dashboard'))
    appt = dbconn.execute('SELECT * FROM appointment WHERE id = ?', (report['appointment_id'],)).fetchone()
    endoscopist = None
    if report['endoscopist_id']:
        endoscopist = dbconn.execute(
            'SELECT * FROM endoscopist WHERE id = ?', (report['endoscopist_id'],)
        ).fetchone()
    images = dbconn.execute(
        'SELECT * FROM ercp_report_image WHERE report_id = ? ORDER BY slot', (report_id,)
    ).fetchall()

    qr_url = url_for('ercp_report_view', appointment_id=report['appointment_id'], _external=True)
    qr_data_uri = generate_qr_data_uri(qr_url)

    assistants_lines = [a.strip() for a in (report['assistants'] or '').split(',') if a.strip()]

    report_dict = dict(report)
    cholangiogram_summary = build_cholangiogram_sentence(report_dict)
    stent_summary = describe_stent_placement(report_dict)
    papilla_summary = describe_papilla_summary(report_dict)

    procedure_fields = [
        ('Procedure', 'ERCP'),
        ('Indication', report['indication']),
        ('Sedation', report['sedation']),
        ('Papilla', papilla_summary),
        ('Cannulation', report['cannulation']),
        ('Cholangiogram Findings', cholangiogram_summary),
        ('Therapeutic Procedures', report['therapeutic_procedures']),
        ('Biliary Stent', stent_summary),
        ('Biopsy', report['biopsy']),
        ('Complications', report['complications']),
    ]
    procedure_fields = [(label, val) for label, val in procedure_fields if val and val.strip()]

    return render_template(
        'ercp_print.html',
        report=report, appt=appt, endoscopist=endoscopist, images=images,
        procedure_fields=procedure_fields, qr_data_uri=qr_data_uri,
        assistants_lines=assistants_lines,
        report_number=f'ERCP-{report_id}',
    )


@app.route('/admin/endoscopists')
@roles_required(*CAN_MANAGE_ENDOSCOPISTS)
def endoscopists_view():
    dbconn = get_db()
    endoscopists = dbconn.execute('SELECT * FROM endoscopist ORDER BY is_active DESC, full_name').fetchall()
    usage_counts = {
        row['endoscopist_id']: row['c']
        for row in dbconn.execute(
            'SELECT endoscopist_id, COUNT(*) as c FROM ercp_report '
            'WHERE endoscopist_id IS NOT NULL GROUP BY endoscopist_id'
        ).fetchall()
    }
    return render_template('endoscopists.html', endoscopists=endoscopists, usage_counts=usage_counts)


@app.route('/admin/endoscopists/add', methods=['POST'])
@roles_required(*CAN_MANAGE_ENDOSCOPISTS)
def add_endoscopist():
    full_name = request.form.get('full_name', '').strip()
    title_lines = request.form.get('title_lines', '').strip()
    if not full_name:
        flash('Please provide the endoscopist\'s name.', 'error')
        return redirect(url_for('endoscopists_view'))
    dbconn = get_db()
    dbconn.execute(
        'INSERT INTO endoscopist (full_name, title_lines, is_active, created_at) VALUES (?,?,1,?)',
        (full_name, title_lines, datetime.utcnow().isoformat())
    )
    dbconn.commit()
    flash(f'Added endoscopist: {full_name}.', 'success')
    return redirect(url_for('endoscopists_view'))


@app.route('/admin/endoscopists/toggle/<int:endoscopist_id>', methods=['POST'])
@roles_required(*CAN_MANAGE_ENDOSCOPISTS)
def toggle_endoscopist(endoscopist_id):
    dbconn = get_db()
    row = dbconn.execute('SELECT * FROM endoscopist WHERE id = ?', (endoscopist_id,)).fetchone()
    if row:
        new_val = 0 if row['is_active'] else 1
        dbconn.execute('UPDATE endoscopist SET is_active = ? WHERE id = ?', (new_val, endoscopist_id))
        dbconn.commit()
    return redirect(url_for('endoscopists_view'))


@app.route('/admin/endoscopists/delete/<int:endoscopist_id>', methods=['POST'])
@roles_required(*CAN_MANAGE_ENDOSCOPISTS)
def delete_endoscopist(endoscopist_id):
    dbconn = get_db()
    in_use = dbconn.execute(
        'SELECT COUNT(*) as c FROM ercp_report WHERE endoscopist_id = ?', (endoscopist_id,)
    ).fetchone()['c']
    if in_use:
        flash('Cannot delete — this endoscopist is referenced by existing reports. Deactivate instead.', 'error')
        return redirect(url_for('endoscopists_view'))
    dbconn.execute('DELETE FROM endoscopist WHERE id = ?', (endoscopist_id,))
    dbconn.commit()
    flash('Endoscopist removed.', 'success')
    return redirect(url_for('endoscopists_view'))


# ----------------------------------------------------------------------
# Bootstrap
# ----------------------------------------------------------------------
init_db()

if __name__ == '__main__':
    app.run(debug=True)
