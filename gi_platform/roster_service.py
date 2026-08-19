"""On-call roster management — separate trainee and house officer schedules."""

from __future__ import annotations

import io
import json
from datetime import date, datetime

from openpyxl import Workbook, load_workbook

from gi_platform.constants import ROSTER_TYPE_LABELS, ROSTER_TYPES, SHIFT_TYPES


def get_or_create_period(db, *, roster_type: str, year_month: str, created_by: int | None) -> int:
    row = db.execute(
        'SELECT id FROM gi_duty_roster_period WHERE roster_type = ? AND year_month = ?',
        (roster_type, year_month),
    ).fetchone()
    if row:
        return row['id']
    title = f"{ROSTER_TYPE_LABELS.get(roster_type, roster_type)} — {year_month}"
    cur = db.execute(
        """
        INSERT INTO gi_duty_roster_period (roster_type, year_month, title, created_by)
        VALUES (?, ?, ?, ?)
        """,
        (roster_type, year_month, title, created_by),
    )
    db.commit()
    return cur.lastrowid


def get_period(db, period_id: int):
    return db.execute('SELECT * FROM gi_duty_roster_period WHERE id = ?', (period_id,)).fetchone()


def list_periods(db, roster_type: str) -> list:
    return db.execute(
        """
        SELECT p.*, u.full_name AS created_by_name
        FROM gi_duty_roster_period p
        LEFT JOIN user u ON u.id = p.created_by
        WHERE p.roster_type = ?
        ORDER BY p.year_month DESC
        """,
        (roster_type,),
    ).fetchall()


def _clear_period_shifts(db, period_id: int) -> None:
    shifts = db.execute('SELECT id FROM gi_duty_roster_shift WHERE period_id = ?', (period_id,)).fetchall()
    for s in shifts:
        db.execute('DELETE FROM gi_duty_roster_assignment WHERE shift_id = ?', (s['id'],))
    db.execute('DELETE FROM gi_duty_roster_shift WHERE period_id = ?', (period_id,))


def set_shift_assignments(
    db, *, period_id: int, roster_date: str, shift_type: str,
    user_ids: list[int], notes: str = '',
) -> None:
    row = db.execute(
        """
        SELECT id FROM gi_duty_roster_shift
        WHERE period_id = ? AND roster_date = ? AND shift_type = ?
        """,
        (period_id, roster_date, shift_type),
    ).fetchone()
    if row:
        shift_id = row['id']
        db.execute('DELETE FROM gi_duty_roster_assignment WHERE shift_id = ?', (shift_id,))
        db.execute(
            'UPDATE gi_duty_roster_shift SET notes = ? WHERE id = ?', (notes, shift_id)
        )
    else:
        cur = db.execute(
            """
            INSERT INTO gi_duty_roster_shift (period_id, roster_date, shift_type, notes)
            VALUES (?, ?, ?, ?)
            """,
            (period_id, roster_date, shift_type, notes),
        )
        shift_id = cur.lastrowid
    for uid in user_ids:
        if uid:
            db.execute(
                'INSERT INTO gi_duty_roster_assignment (shift_id, user_id) VALUES (?, ?)',
                (shift_id, uid),
            )
    db.commit()


def get_period_grid(db, period_id: int) -> dict:
    period = get_period(db, period_id)
    if not period:
        return {}
    shifts = db.execute(
        'SELECT * FROM gi_duty_roster_shift WHERE period_id = ? ORDER BY roster_date, shift_type',
        (period_id,),
    ).fetchall()
    grid = {}
    for sh in shifts:
        assigns = db.execute(
            """
            SELECT a.user_id, u.full_name
            FROM gi_duty_roster_assignment a
            JOIN user u ON u.id = a.user_id
            WHERE a.shift_id = ?
            ORDER BY u.full_name
            """,
            (sh['id'],),
        ).fetchall()
        key = (sh['roster_date'], sh['shift_type'])
        grid[key] = {'shift': sh, 'assignments': assigns}
    return {'period': period, 'grid': grid}


def publish_period(db, period_id: int, published_by: int) -> list[int]:
    """Publish roster and notify all assigned users. Returns notified user IDs."""
    from gi_platform import notification_service

    period = get_period(db, period_id)
    if not period:
        return []
    db.execute(
        """
        UPDATE gi_duty_roster_period
        SET status = 'published', published_by = ?, published_at = datetime('now')
        WHERE id = ?
        """,
        (published_by, period_id),
    )
    db.commit()

    user_shifts: dict[int, list[str]] = {}
    grid = get_period_grid(db, period_id)
    for (rdate, stype), cell in grid.get('grid', {}).items():
        label = f"{rdate} ({stype})"
        for a in cell['assignments']:
            user_shifts.setdefault(a['user_id'], []).append(label)

    notified = []
    title = ROSTER_TYPE_LABELS.get(period['roster_type'], 'On-call roster')
    for uid, dates in user_shifts.items():
        body = f"You are scheduled for {title} {period['year_month']}:\n" + '\n'.join(dates)
        notification_service.notify_user(
            db, user_id=uid,
            title=f'{title} published',
            body=body,
            link_url=f'/roster/my-duties',
        )
        notified.append(uid)
    db.commit()
    return notified


def my_duties(db, user_id: int, *, from_date: str | None = None) -> list:
    sql = """
        SELECT p.roster_type, p.year_month, p.title, s.roster_date, s.shift_type, s.notes
        FROM gi_duty_roster_assignment a
        JOIN gi_duty_roster_shift s ON s.id = a.shift_id
        JOIN gi_duty_roster_period p ON p.id = s.period_id
        WHERE a.user_id = ? AND p.status = 'published'
    """
    params: list = [user_id]
    if from_date:
        sql += ' AND s.roster_date >= ?'
        params.append(from_date)
    sql += ' ORDER BY s.roster_date'
    return db.execute(sql, params).fetchall()


def import_excel(db, *, period_id: int, file_bytes: bytes, role_filter: str) -> int:
    """Import roster from Excel: columns Date, Shift, Doctor(s) comma-separated usernames."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    count = 0
    users = {r['username'].lower(): r['id'] for r in db.execute('SELECT id, username FROM user').fetchall()}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not row or not row[0]:
            continue
        rdate = str(row[0])[:10]
        shift_type = (str(row[1]) if len(row) > 1 and row[1] else 'on_call').strip().lower()
        if shift_type not in SHIFT_TYPES:
            shift_type = 'on_call'
        names = str(row[2] if len(row) > 2 else '').split(',')
        uids = []
        for n in names:
            uname = n.strip().lower()
            if uname in users:
                uids.append(users[uname])
        if uids:
            set_shift_assignments(db, period_id=period_id, roster_date=rdate,
                                  shift_type=shift_type, user_ids=uids)
            count += 1
    return count


def export_excel(db, period_id: int) -> bytes:
    grid = get_period_grid(db, period_id)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Roster'
    ws.append(['Date', 'Shift', 'Doctors'])
    for (rdate, stype), cell in sorted(grid.get('grid', {}).items()):
        names = ', '.join(a['full_name'] for a in cell['assignments'])
        ws.append([rdate, stype, names])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def eligible_staff(db, roster_type: str) -> list:
    role = 'pg_trainee' if roster_type == 'pg_trainee' else 'house_officer'
    return db.execute(
        "SELECT id, full_name, username FROM user WHERE role = ? AND is_approved = 1 ORDER BY full_name",
        (role,),
    ).fetchall()
