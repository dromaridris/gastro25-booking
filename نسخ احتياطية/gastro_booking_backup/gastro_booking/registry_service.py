"""
Gastro25 Core Services — Registry Service
--------------------------------------------
Generic, reusable infrastructure for listing/searching/exporting
procedure records (appointments, reports, etc.) — intended as the
foundation for every future procedure registry (ERCP, Dilatation, PEG,
APC, EUS, ...).

READ BEFORE WIRING THIS INTO A NEW PAGE:
This application does not currently have a dedicated, browsable "ERCP
Registry" page (with in-browser search/pagination/sorting). The closest
existing feature is the Admin "Export Bookings to Excel" date-range
export. This phase extracts the reusable pieces of that feature (dynamic
column definitions, generic styled Excel export) into this service and
refactors the existing export route to use it — with no change to its
visible output or workflow, per the instruction not to modify the
current workflow.

The filter/search/sort/paginate helpers below are provided so a future
in-browser Registry page (for ERCP or any new procedure) doesn't have to
reinvent them — but no such page exists yet, and adding one is a
separate, explicitly-scoped UI change, not part of this phase.

Design: a "registry definition" for a procedure is just:
    - a base SQL query (owned by the procedure module, e.g. app.py)
    - a list of Column(key, label, formatter) — dynamic columns
    - optional filters (date range, search term)
This service never contains SQL specific to any one procedure — all rows
and SQL text are supplied by the caller; this module only orders,
filters, paginates, and renders/export them.
"""

from dataclasses import dataclass
from typing import Any, Callable, Optional


@dataclass
class Column:
    """One dynamic column in a registry listing/export.

    key:       the field name to read the raw value from (dict or
               sqlite3.Row).
    label:     human-facing column header.
    formatter: optional callable(raw_value, row) -> display value, e.g.
               a procedure-label lookup, a Yes/No flag, a role label.
               Defaults to the raw value when omitted.
    width:     optional Excel column width. Falls back to a size derived
               from the label when omitted.
    """
    key: str
    label: str
    formatter: Optional[Callable[[Any, Any], Any]] = None
    width: Optional[int] = None

    def value_for(self, row):
        raw = row[self.key] if self.key in row.keys() else None
        return self.formatter(raw, row) if self.formatter else raw


def filter_by_date_range(rows, date_key, start_iso, end_iso):
    """Generic inclusive date-range filter over an in-memory row list —
    for use when the base query doesn't already scope by date itself."""
    return [r for r in rows if start_iso <= r[date_key] <= end_iso]


def search_rows(rows, columns, term):
    """Generic case-insensitive substring search across a set of
    columns' *display* values (so it matches what the user actually
    sees, e.g. the procedure label rather than the raw DB code)."""
    if not term:
        return rows
    term = term.lower()
    matched = []
    for row in rows:
        for col in columns:
            val = col.value_for(row)
            if val is not None and term in str(val).lower():
                matched.append(row)
                break
    return matched


def sort_rows(rows, sort_key, descending=False):
    """Generic sort by any row key. Rows missing that key (or with a
    None value) sort last regardless of direction."""
    def keyfn(row):
        val = row[sort_key] if sort_key in row.keys() else None
        return (val is None, val)
    return sorted(rows, key=keyfn, reverse=descending)


def paginate(rows, page, per_page):
    """Generic 1-indexed pagination.
    Returns (page_rows, total_count, total_pages)."""
    total = len(rows)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    start = (page - 1) * per_page
    return rows[start:start + per_page], total, total_pages


def build_excel_workbook(rows, columns, sheet_title='Registry', header_fill_color='A6192E'):
    """Generic Excel export: any row list + Column definitions -> a
    styled openpyxl Workbook (bold white header text on the department's
    crimson fill, centered header, frozen header row, auto-filter, and
    per-column widths) — styling identical to the original bookings
    export this replaces."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    headers = [col.label for col in columns]
    ws.append(headers)
    header_fill = PatternFill('solid', start_color=header_fill_color, end_color=header_fill_color)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row in rows:
        ws.append([col.value_for(row) for col in columns])

    for i, col in enumerate(columns, start=1):
        width = col.width if col.width is not None else max(12, min(34, len(col.label) + 4))
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{max(ws.max_row, 1)}'

    return wb
